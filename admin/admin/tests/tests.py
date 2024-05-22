# SPDX-FileCopyrightText: 2023 Magenta ApS <info@magenta.dk>
#
# SPDX-License-Identifier: MPL-2.0
import base64
import csv
import json
import re
import time
import traceback
from collections import defaultdict
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Dict, List, Tuple
from unittest.mock import mock_open, patch
from urllib.parse import parse_qs, quote, quote_plus, urlparse

import requests
from bs4 import BeautifulSoup, Tag
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.test import TestCase
from django.urls import reverse
from django.views.generic import TemplateView
from openpyxl import Workbook, load_workbook
from requests import Response
from told_common.data import Notat, Vareafgiftssats
from told_common.forms import TF10VareForm
from told_common.rest_client import RestClient
from told_common.tests import (
    AnmeldelseListViewTest,
    HasLogin,
    LoginTest,
    PermissionsTest,
    TemplateTagsTest,
    TestMixin,
    modify_values,
)
from told_common.views import FragtbrevView

from admin.clients.prisme import PrismeClient, prisme_send_dummy
from admin.forms import TF10CreateForm
from admin.views import (
    AfgiftstabelDetailView,
    AfgiftstabelDownloadView,
    AfgiftstabelListView,
    StatistikView,
    TF5View,
    TF10EditMultipleView,
    TF10FormUpdateView,
    TF10HistoryDetailView,
    TF10HistoryListView,
    TF10ListView,
    TF10View,
)


class TestLogin(TestMixin, TestCase):
    @staticmethod
    def create_response(status_code, content):
        response = Response()
        response.status_code = status_code
        if type(content) in (dict, list):
            content = json.dumps(content)
        if type(content) is str:
            content = content.encode("utf-8")
        response._content = content
        return response

    def submit_get_errors(self, url, data):
        return self.get_errors(self.client.post(url, data=data).content)

    @classmethod
    def mock_requests_get(cls, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        status_code = None
        json_content = None
        content = None
        if path in (expected_prefix + "user", expected_prefix + "user/this"):
            json_content = {
                "id": 1,
                "username": "admin",
                "first_name": "Administrator",
                "last_name": "",
                "email": "admin@told.gl",
                "is_superuser": True,
                "groups": [],
                "permissions": [],
            }
        elif path == expected_prefix + "afsender":
            json_content = {"count": 0, "items": []}
        else:
            raise Exception(f"Mock {cls.__name__} got unrecognized path: GET {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests, "post", return_value=create_response(401, "Unauthorized"))
    def test_incorrect_login(self, mock_post, mock_get):
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(reverse("login"), {"username": "incorrect"})
        self.assertEquals(response.status_code, 200)  # Rerender form
        mock_post.assert_not_called()
        errors = self.get_errors(response.content)
        self.assertEquals(errors["password"], ["Dette felt er påkrævet."])

        response = self.client.post(
            reverse("login"), {"username": "incorrect", "password": "credentials"}
        )
        self.assertEquals(response.status_code, 200)  # Rerender form
        mock_post.assert_called_with(
            f"{settings.REST_DOMAIN}/api/token/pair",
            json={"username": "incorrect", "password": "credentials"},
            headers={"Content-Type": "application/json"},
        )

    @patch.object(requests.sessions.Session, "get")
    @patch.object(
        requests,
        "post",
        return_value=create_response(200, {"access": "123456", "refresh": "abcdef"}),
    )
    def test_correct_login(self, mock_post, mock_get):
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(
            reverse("login") + "?back=/",
            {"username": "correct", "password": "credentials"},
        )
        mock_post.assert_called_with(
            f"{settings.REST_DOMAIN}/api/token/pair",
            json={"username": "correct", "password": "credentials"},
            headers={"Content-Type": "application/json"},
        )
        self.assertEquals(response.status_code, 302)
        self.assertEquals(response.headers["Location"], "/")
        self.assertEquals(self.client.session["access_token"], "123456")
        self.assertEquals(self.client.session["refresh_token"], "abcdef")

    @patch.object(RestClient, "refresh_login")
    @patch.object(
        requests.sessions.Session,
        "get",
    )
    @patch.object(
        requests,
        "post",
        return_value=create_response(200, {"access": "123456", "refresh": "abcdef"}),
    )
    def test_token_refresh_not_needed(self, mock_post, mock_get, mock_refresh_login):
        mock_get.side_effect = self.mock_requests_get
        self.client.post(
            reverse("login"), {"username": "correct", "password": "credentials"}
        )
        self.client.get(reverse("rest", kwargs={"path": "afsender"}))
        # Check that token refresh is not needed
        mock_refresh_login.assert_not_called()
        with self.settings(NINJA_JWT={"ACCESS_TOKEN_LIFETIME": timedelta(seconds=1)}):
            self.client.get(reverse("rest", kwargs={"path": "afsender"}))
            # Check that token refresh is needed
            mock_refresh_login.assert_called()

    @patch.object(RestClient, "refresh_login")
    @patch.object(
        requests.sessions.Session,
        "get",
    )
    @patch.object(
        requests,
        "post",
        return_value=create_response(200, {"access": "123456", "refresh": "abcdef"}),
    )
    def test_token_refresh_needed(self, mock_post, mock_get, mock_refresh_login):
        mock_get.side_effect = self.mock_requests_get
        self.client.post(
            reverse("login"), {"username": "correct", "password": "credentials"}
        )
        self.client.get(reverse("rest", kwargs={"path": "afsender"}))
        # Check that token refresh is not needed
        mock_refresh_login.assert_not_called()
        # Set token max_age way down, so it will be refreshed
        with self.settings(NINJA_JWT={"ACCESS_TOKEN_LIFETIME": timedelta(seconds=1)}):
            self.client.get(reverse("rest", kwargs={"path": "afsender"}))
            # Check that token refresh is needed
            mock_refresh_login.assert_called()

    @patch.object(
        requests.sessions.Session,
        "get",
    )
    @patch.object(
        requests,
        "post",
    )
    def test_token_refresh(self, mock_post, mock_get):
        mock_get.side_effect = self.mock_requests_get
        mock_post.return_value = self.create_response(
            200, {"access": "123456", "refresh": "abcdef"}
        )
        self.client.post(
            reverse("login"), {"username": "correct", "password": "credentials"}
        )
        mock_post.return_value = self.create_response(200, {"access": "7890ab"})
        # Set token max_age way down, so it will be refreshed
        with self.settings(NINJA_JWT={"ACCESS_TOKEN_LIFETIME": timedelta(seconds=1)}):
            response = self.client.get(reverse("rest", kwargs={"path": "afsender"}))
            # Check that token refresh is needed
            self.assertEquals(self.client.session["access_token"], "7890ab")

    @patch.object(requests.sessions.Session, "get")
    @patch.object(
        requests,
        "post",
        return_value=create_response(200, {"access": "123456", "refresh": "abcdef"}),
    )
    def test_logout(self, mock_post, mock_get):
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(
            reverse("login") + "?back=/",
            {"username": "correct", "password": "credentials"},
        )
        self.assertEquals(response.status_code, 302)
        self.assertEquals(response.headers["Location"], "/")
        self.assertEquals(self.client.session["access_token"], "123456")
        self.assertEquals(self.client.session["refresh_token"], "abcdef")
        response = self.client.get(reverse("logout"))
        self.assertNotIn("access_token", self.client.session)
        self.assertNotIn("refresh_token", self.client.session)

    def test_token_refresh_expired(self):
        session = self.client.session
        session.update(
            {
                "access_token": "123456",
                "refresh_token": "abcdef",
                "access_token_timestamp": time.time(),
                "refresh_token_timestamp": (
                    datetime.now() - timedelta(days=2)
                ).timestamp(),
            }
        )
        session.save()
        session_cookie = settings.SESSION_COOKIE_NAME
        self.client.cookies[session_cookie] = session.session_key
        cookie_data = {
            "max-age": None,
            "path": "/admin/",
            "domain": settings.SESSION_COOKIE_DOMAIN,
            "secure": settings.SESSION_COOKIE_SECURE or None,
            "expires": None,
        }
        self.client.cookies[session_cookie].update(cookie_data)

        response = self.client.get(reverse("tf10_list"))
        self.assertEquals(response.status_code, 302)
        self.assertEquals(
            response.headers["Location"],
            "/admin/login?back=" + quote_plus(reverse("tf10_list")),
        )


class TestGodkend(PermissionsTest, TestCase):
    view = TF10View

    check_permissions = (
        (reverse("tf10_view", kwargs={"id": 1}), view.required_permissions),
    )

    def setUp(self):
        super().setUp()
        self.patched: List[Tuple[str, str]] = []
        self.posted: List[Tuple[str, str]] = []

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "vareafgiftssats/1":
            json_content = {
                "id": 1,
                "afgiftstabel": 1,
                "vareart_da": "Båthorn",
                "vareart_kl": "Båthorn",
                "afgiftsgruppenummer": 1234567,
                "enhed": "kg",
                "afgiftssats": "1.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            }

        elif path == expected_prefix + "afsender/1":
            json_content = {
                "id": 1,
                "navn": "Testfirma 1",
                "adresse": "Testvej 42",
                "postnummer": 1234,
                "by": "TestBy",
                "postbox": "123",
                "telefon": "123456",
                "cvr": 12345678,
            }

        elif path == expected_prefix + "modtager/1":
            json_content = {
                "id": 1,
                "navn": "Testfirma 1",
                "adresse": "Testvej 42",
                "postnummer": 1234,
                "by": "TestBy",
                "postbox": "123",
                "telefon": "123456",
                "cvr": 12345678,
                "kreditordning": True,
                "stedkode": 123,
            }

        elif path == expected_prefix + "afgiftsanmeldelse/1":
            json_content = {
                "id": 1,
                "afsender": 1,
                "modtager": 1,
                "fragtforsendelse": 1,
                "postforsendelse": None,
                "leverandørfaktura_nummer": "1234",
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "betales_af": "afsender",
                "indførselstilladelse": "5678",
                "afgift_total": "5000.00",
                "betalt": False,
                "dato": "2023-08-22T00:00:00-02:00",
                "status": "afvist",
                "beregnet_faktureringsdato": "2023-10-10",
            }
        elif path == expected_prefix + "afgiftsanmeldelse/1/full":
            json_content = {
                "id": 1,
                "afsender": {
                    "id": 1,
                    "navn": "Testfirma 1",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                },
                "modtager": {
                    "id": 1,
                    "navn": "Testfirma 1",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                    "kreditordning": True,
                },
                "fragtforsendelse": {
                    "id": 1,
                    "forsendelsestype": "S",
                    "fragtbrevsnummer": "ABCDE1234567",
                    "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                    "forbindelsesnr": "ABC 337",
                    "afgangsdato": "2023-10-01",
                    "kladde": False,
                },
                "postforsendelse": None,
                "leverandørfaktura_nummer": "1234",
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "betales_af": "afsender",
                "indførselstilladelse": "5678",
                "afgift_total": "5000.00",
                "betalt": False,
                "dato": "2023-08-22T00:00:00-02:00",
                "status": "afvist",
                "beregnet_faktureringsdato": "2023-10-10",
                "oprettet_af": {
                    "username": "testuser1",
                    "first_name": "",
                    "last_name": "",
                    "email": "",
                    "is_superuser": False,
                    "groups": [],
                    "permissions": [
                        "anmeldelse.add_afgiftsanmeldelse",
                        "anmeldelse.change_afgiftsanmeldelse",
                        "anmeldelse.delete_afgiftsanmeldelse",
                        "anmeldelse.view_afgiftsanmeldelse",
                        "anmeldelse.view_all_anmeldelse",
                        "forsendelse.view_all_fragtforsendelser",
                        "forsendelse.view_all_postforsendelser",
                    ],
                    "indberetter_data": {
                        "cpr": 1234567890,
                        "cvr": 12345678,
                    },
                    "twofactor_enabled": False,
                },
            }
        elif path == expected_prefix + "afgiftsanmeldelse/2/full":
            status_code = 404
        elif path == expected_prefix + "fragtforsendelse/1":
            json_content = {
                "id": 1,
                "forsendelsestype": "S",
                "fragtbrevsnummer": "ABCDE1234567",
                "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "forbindelsesnr": "ABC 337",
                "afgangsdato": "2023-10-01",
            }

        elif path == expected_prefix + "varelinje":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "vareafgiftssats": 1,
                        "antal": 5,
                        "mængde": "1.00",
                        "fakturabeløb": "25000.00",
                        "afgiftsbeløb": "5000.00",
                    }
                ],
            }

        elif path == expected_prefix + "notat":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "privatafgiftsanmeldelse": None,
                        "tekst": "Hephey",
                        "navn": "tester",
                        "oprettet": "2023-10-01T00:00:00.000000+00:00",
                        "index": 0,
                    }
                ],
            }
        elif path == expected_prefix + "prismeresponse":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "rec_id": "5637147578",
                        "tax_notification_number": "44668899",
                        "delivery_date": "2023-04-07T00:00:00",
                    }
                ],
            }
        elif path == expected_prefix + "toldkategori":
            json_content = [
                {
                    "kategori": "70",
                    "navn": "RAL Royal Arctic Line A/S",
                    "kræver_cvr": False,
                }
            ]
        else:
            raise Exception(
                f"Mock {self.__class__.__name__} got unrecognized path: GET {path}"
            )
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_patch(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftsanmeldelse/1":
            json_content = {"id": 1}
            self.patched.append((path, data))
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: PATCH {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_post(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "eboks":
            json_content = {"id": 1}
            self.posted.append((path, data))
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: POST {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_error(self, path, *args, **kwargs):
        response = Response()
        response.status_code = 500
        return response

    def mock_requests_error_401(self, path, *args, **kwargs):
        response = Response()
        response.status_code = 401
        return response

    def test_requires_login(self):
        url = str(reverse("tf10_list"))
        response = self.client.get(url)
        self.assertEquals(response.status_code, 302)
        self.assertEquals(
            response.headers["Location"],
            reverse("login") + "?back=" + quote(url, safe=""),
        )

    @patch.object(requests.sessions.Session, "get")
    def test_get_view(self, mock_get):
        self.login()
        url = reverse("tf10_view", kwargs={"id": 1})
        mock_get.side_effect = self.mock_requests_get
        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)

    @patch.object(requests.sessions.Session, "get")
    def test_get_view_not_found(self, mock_get):
        self.login()
        url = reverse("tf10_view", kwargs={"id": 2})
        mock_get.side_effect = self.mock_requests_get
        response = self.client.get(url)
        self.assertEquals(response.status_code, 404)

    @patch.object(requests.sessions.Session, "patch")
    @patch.object(requests.sessions.Session, "get")
    def test_post_view_godkend(self, mock_get, mock_patch):
        self.login()
        view_url = reverse("tf10_view", kwargs={"id": 1})
        mock_patch.side_effect = self.mock_requests_patch
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(view_url, {"status": "godkendt"})
        self.assertEquals(response.status_code, 302)
        prefix = f"{settings.REST_DOMAIN}/api/"
        patched_map = defaultdict(list)
        for url, data in self.patched:
            patched_map[url].append(json.loads(data))
        self.assertEquals(
            patched_map[prefix + "afgiftsanmeldelse/1"], [{"status": "godkendt"}]
        )

    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests.sessions.Session, "post")
    @patch.object(requests.sessions.Session, "patch")
    def test_post_view_afvis(self, mock_patch, mock_post, mock_get):
        self.login()
        view_url = reverse("tf10_view", kwargs={"id": 1})
        mock_patch.side_effect = self.mock_requests_patch
        mock_post.side_effect = self.mock_requests_post
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(view_url, {"status": "afvist"})
        self.assertEquals(response.status_code, 302)
        prefix = f"{settings.REST_DOMAIN}/api/"
        patched_map = defaultdict(list)
        for url, data in self.patched:
            patched_map[url].append(json.loads(data))
        self.assertEquals(
            patched_map[prefix + "afgiftsanmeldelse/1"], [{"status": "afvist"}]
        )

    @patch.object(requests.sessions.Session, "patch")
    @patch.object(requests.sessions.Session, "get")
    def test_post_view_not_found(self, mock_get, mock_patch):
        self.login()
        view_url = reverse("tf10_view", kwargs={"id": 2})
        mock_patch.side_effect = self.mock_requests_patch
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(view_url, {"status": "godkendt"})
        self.assertEquals(response.status_code, 404)
        prefix = f"{settings.REST_DOMAIN}/api/"
        patched_map = defaultdict(list)
        for url, data in self.patched:
            patched_map[url].append(json.loads(data))
        self.assertEquals(patched_map[prefix + "afgiftsanmeldelse/2"], [])

    @patch.object(requests.sessions.Session, "get")
    def test_get_view_rest_error(self, mock_get):
        self.login()
        url = reverse("tf10_view", kwargs={"id": 1})
        mock_get.side_effect = self.mock_requests_error
        response = self.client.get(url)
        self.assertEquals(response.status_code, 500)

    @patch.object(requests.sessions.Session, "get")
    def test_get_view_rest_error_401(self, mock_get):
        self.login()
        url = reverse("tf10_view", kwargs={"id": 1})
        mock_get.side_effect = self.mock_requests_error_401
        response = self.client.get(url)
        self.assertEquals(response.status_code, 302)

    @patch.object(requests.sessions.Session, "patch")
    @patch.object(requests.sessions.Session, "get")
    def test_post_view_rest_error(self, mock_get, mock_patch):
        self.login()
        view_url = reverse("tf10_view", kwargs={"id": 1})
        mock_get.side_effect = self.mock_requests_get
        mock_patch.side_effect = self.mock_requests_error
        response = self.client.post(view_url, {"status": "godkendt"})
        self.assertEquals(response.status_code, 500)


class TestPrisme(TestMixin, PermissionsTest, TestCase):
    view = TF10View

    check_permissions = (
        (reverse("tf10_view", kwargs={"id": 1}), view.required_permissions),
    )

    def setUp(self):
        super().setUp()
        self.posted: List[Tuple[str, str]] = []
        self.patched: List[Tuple[str, str]] = []

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "vareafgiftssats/1":
            json_content = {
                "id": 1,
                "afgiftstabel": 1,
                "vareart_da": "Båthorn",
                "vareart_kl": "Båthorn",
                "afgiftsgruppenummer": 1234567,
                "enhed": "kg",
                "afgiftssats": "1.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            }

        elif path == expected_prefix + "afsender/1":
            json_content = {
                "id": 1,
                "navn": "Testfirma 1",
                "adresse": "Testvej 42",
                "postnummer": 1234,
                "by": "TestBy",
                "postbox": "123",
                "telefon": "123456",
                "cvr": 12345678,
            }

        elif path == expected_prefix + "modtager/1":
            json_content = {
                "id": 1,
                "navn": "Testfirma 1",
                "adresse": "Testvej 42",
                "postnummer": 1234,
                "by": "TestBy",
                "postbox": "123",
                "telefon": "123456",
                "cvr": 12345678,
                "kreditordning": True,
            }

        elif path == expected_prefix + "afgiftsanmeldelse/1":
            json_content = {
                "id": 1,
                "afsender": 1,
                "modtager": 1,
                "fragtforsendelse": 1,
                "postforsendelse": None,
                "leverandørfaktura_nummer": "1234",
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "betales_af": "afsender",
                "indførselstilladelse": "5678",
                "afgift_total": "5000.00",
                "betalt": False,
                "dato": "2023-08-22T00:00:00-02:00",
                "status": "afvist",
                "beregnet_faktureringsdato": "2023-10-10",
            }
        elif path == expected_prefix + "afgiftsanmeldelse/1/full":
            json_content = {
                "id": 1,
                "afsender": {
                    "id": 1,
                    "navn": "Testfirma 1",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                },
                "modtager": {
                    "id": 1,
                    "navn": "Testfirma 1",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                    "kreditordning": True,
                },
                "fragtforsendelse": {
                    "id": 1,
                    "forsendelsestype": "S",
                    "fragtbrevsnummer": "ABCDE1234567",
                    "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                    "forbindelsesnr": "ABC 337",
                    "afgangsdato": "2023-10-01",
                },
                "postforsendelse": None,
                "leverandørfaktura_nummer": "1234",
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "betales_af": "afsender",
                "indførselstilladelse": "5678",
                "afgift_total": "5000.00",
                "betalt": False,
                "dato": "2023-08-22T00:00:00-02:00",
                "status": "afvist",
                "beregnet_faktureringsdato": "2023-10-10",
            }
        elif path == expected_prefix + "fragtforsendelse/1":
            json_content = {
                "id": 1,
                "forsendelsestype": "S",
                "fragtbrevsnummer": "ABCDE1234567",
                "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "forbindelsesnr": "ABC 337",
                "afgangsdato": "2023-10-01",
            }

        elif path == expected_prefix + "varelinje":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "vareafgiftssats": 1,
                        "antal": 5,
                        "mængde": "1.00",
                        "fakturabeløb": "25000.00",
                        "afgiftsbeløb": "5000.00",
                    }
                ],
            }

        elif path == expected_prefix + "notat":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "privatafgiftsanmeldelse": None,
                        "tekst": "Hephey",
                        "navn": "tester",
                        "oprettet": "2023-10-01T00:00:00.000000+00:00",
                        "index": 0,
                    }
                ],
            }
        elif path == expected_prefix + "prismeresponse":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "rec_id": "5637147578",
                        "tax_notification_number": "44668899",
                        "delivery_date": "2023-04-07T00:00:00",
                    }
                ],
            }
        elif path == expected_prefix + "toldkategori":
            json_content = [
                {
                    "kategori": "70",
                    "navn": "RAL Royal Arctic Line A/S",
                    "kræver_cvr": False,
                }
            ]
        else:
            raise Exception(
                f"Mock {self.__class__.__name__} got unrecognized path: {path}"
            )
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_post(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "prismeresponse":
            json_content = {"id": 1}
            self.posted.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_patch(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "modtager/1":
            json_content = {"id": 1}
            self.patched.append((path, data))
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: PATCH {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_error(self, path, *args, **kwargs):
        response = Response()
        response.status_code = 500
        return response

    def mock_requests_error_401(self, path, *args, **kwargs):
        response = Response()
        response.status_code = 401
        return response

    def mock_prisme_send(self, request_object):
        return prisme_send_dummy(request_object)

    @patch.object(requests.sessions.Session, "patch")
    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests.sessions.Session, "post")
    @patch.object(PrismeClient, "send")
    def test_post_view_prisme(self, mock_send, mock_post, mock_get, mock_patch):
        self.login()
        view_url = reverse("tf10_view", kwargs={"id": 1})
        mock_send.side_effect = self.mock_prisme_send
        mock_post.side_effect = self.mock_requests_post
        mock_get.side_effect = self.mock_requests_get
        mock_patch.side_effect = self.mock_requests_patch
        response = self.client.post(view_url, {"send_til_prisme": "true"})
        self.assertEquals(response.status_code, 200)
        self.assertEquals(
            self.get_errors(response.content)["__all__"],
            ["Der skal vælges en stedkode når der sendes til Prisme"],
        )

        response = self.client.post(
            view_url, {"send_til_prisme": "true", "modtager_stedkode": "123"}
        )
        self.assertEquals(response.status_code, 302)
        prefix = f"{settings.REST_DOMAIN}/api/"
        patched = defaultdict(list)
        for url, data in self.patched:
            patched[url].append(json.loads(data))
        self.assertEquals(patched[prefix + "modtager/1"][0]["stedkode"], 123)
        posted = defaultdict(list)
        for url, data in self.posted:
            posted[url].append(json.loads(data))
        self.assertEquals(
            posted[prefix + "prismeresponse"],
            [
                {
                    "afgiftsanmeldelse_id": 1,
                    "delivery_date": "2023-04-07T00:00:00",
                    "rec_id": "5637147578",
                    "tax_notification_number": "44668899",
                }
            ],
        )


class FileViewTest(PermissionsTest, TestCase):
    view = FragtbrevView

    @property
    def login_url(self):
        return str(reverse("login"))

    check_permissions = (
        (reverse("fragtbrev_view", kwargs={"id": 1}), view.required_permissions),
    )

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "fragtforsendelse/1":
            json_content = {
                "id": 1,
                "forsendelsestype": "S",
                "fragtbrevsnummer": "ABCDE1234567",
                "fragtbrev": "/fragtbreve/1/fragtbrev.txt",
                "forbindelsesnr": "ABC 337",
                "afgangsdato": "2023-10-01",
            }
        if path == expected_prefix + "fragtforsendelse/2":
            json_content = {
                "id": 2,
                "forsendelsestype": "S",
                "fragtbrevsnummer": 2,
                "fragtbrev": None,
            }
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response


class AdminTemplateTagsTest(TemplateTagsTest, TestCase):
    pass


class AdminLoginTest(LoginTest, TestCase):
    @property
    def restricted_url(self):
        return str(reverse("tf10_list"))


class AdminAnmeldelseListViewTest(PermissionsTest, AnmeldelseListViewTest, TestCase):
    can_view = True
    can_edit = True
    can_select_multiple = True
    view = TF10ListView
    check_permissions = ((reverse("tf10_list"), view.required_permissions),)

    @property
    def login_url(self):
        return str(reverse("login"))

    @property
    def list_url(self):
        return str(reverse("tf10_list"))

    def edit_url(self, id: int):
        return str(reverse("tf10_edit", kwargs={"id": id}))

    def view_url(self, id: int):
        return str(reverse("tf10_view", kwargs={"id": id}))


class AnmeldelseHistoryListViewTest(PermissionsTest, TestCase):
    view = TF10HistoryListView
    check_permissions = (
        (reverse("tf10_history", kwargs={"id": 1}), view.required_permissions),
    )

    @property
    def login_url(self):
        return str(reverse("login"))

    @staticmethod
    def get_html_list(html: str):
        soup = BeautifulSoup(html, "html.parser")
        headers = [element.text for element in soup.css.select("table thead tr th")]
        return [
            dict(zip(headers, [td.text.strip() for td in row.select("td")]))
            for row in soup.css.select("table tbody tr")
        ]

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.testdata = [
            {
                "id": 1,
                "leverandørfaktura_nummer": "12345",
                "betales_af": "afsender",
                "indførselstilladelse": "abcde",
                "betalt": False,
                "leverandørfaktura": "/leverand%C3%B8rfakturaer"
                "/10/leverand%C3%B8rfaktura.pdf",
                "afsender": 1,
                "modtager": 1,
                "postforsendelse": 1,
                "dato": "2023-09-03T00:00:00-02:00",
                "beregnet_faktureringsdato": "2023-10-10",
                "afgift_total": "0.00",
                "fragtforsendelse": None,
                "status": "afvist",
                "history_username": "admin",
                "history_date": "2023-10-01",
            },
            {
                "id": 1,
                "leverandørfaktura_nummer": "12345",
                "betales_af": "afsender",
                "indførselstilladelse": "abcde",
                "betalt": False,
                "leverandørfaktura": "/leverand%C3%B8rfakturaer"
                "/10/leverand%C3%B8rfaktura.pdf",
                "afsender": 1,
                "modtager": 1,
                "postforsendelse": 1,
                "dato": "2023-09-03T00:00:00-02:00",
                "beregnet_faktureringsdato": "2023-10-10",
                "afgift_total": "0.00",
                "fragtforsendelse": None,
                "status": "godkendt",
                "history_username": "admin",
                "history_date": "2023-10-02",
            },
        ]

    def mock_requests_get(self, path):
        expected_prefix = "/api/"
        p = urlparse(path)
        path = p.path
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftsanmeldelse/1/history":
            items = deepcopy(self.testdata)
            json_content = {"count": len(items), "items": items}
        elif path == expected_prefix + "notat":
            json_content = {"count": 0, "items": []}
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    def test_list_view(self, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        url = reverse("tf10_history", kwargs={"id": 1})
        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)
        list_items = self.get_html_list(response.content)
        self.assertEquals(
            list_items,
            [
                {
                    "Dato": "2023-10-01 00:00:00",
                    "Ændret af": "admin",
                    "Notat": "",
                    "Handlinger": "Vis",
                },
                {
                    "Dato": "2023-10-02 00:00:00",
                    "Ændret af": "admin",
                    "Notat": "",
                    "Handlinger": "Vis",
                },
            ],
        )


class AnmeldelseHistoryDetailViewTest(PermissionsTest, TestCase):
    view = TF10HistoryDetailView
    check_permissions = (
        (
            reverse("tf10_history_view", kwargs={"id": 1, "index": 0}),
            view.required_permissions,
        ),
    )

    @property
    def login_url(self):
        return str(reverse("login"))

    def mock_requests_get(self, path):
        expected_prefix = "/api/"
        p = urlparse(path)
        path = p.path
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftsanmeldelse/1/history/0":
            json_content = {
                "id": 1,
                "afsender": {
                    "id": 1,
                    "navn": "Testfirma1",
                    "adresse": "Testvej 1",
                    "postnummer": 1234,
                    "by": "Testby",
                    "postbox": None,
                    "telefon": "123456",
                    "cvr": 12345678,
                },
                "modtager": {
                    "id": 2,
                    "navn": "Testfirma2",
                    "adresse": "Testvej 2",
                    "postnummer": 1234,
                    "by": "Testby",
                    "postbox": None,
                    "telefon": "789012",
                    "cvr": 12345679,
                    "kreditordning": False,
                },
                "fragtforsendelse": {
                    "id": 1,
                    "forsendelsestype": "F",
                    "fragtbrevsnummer": "0",
                    "fragtbrev": "/fragtbreve/1/fragtbrev.txt",
                    "forbindelsesnr": "7331",
                    "afgangsdato": "2023-10-01",
                },
                "postforsendelse": None,
                "leverandørfaktura_nummer": "1234",
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "betales_af": "afsender",
                "indførselstilladelse": "5678",
                "afgift_total": "5000.00",
                "betalt": False,
                "dato": "2023-10-06T00:00:00-02:00",
                "beregnet_faktureringsdato": "2023-10-10",
                "status": "ny",
                "history_username": "admin",
                "history_date": "2023-10-01T00:00:00.000000+00:00",
            }
        elif path == expected_prefix + "varelinje":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "vareafgiftssats": 1,
                        "antal": 5,
                        "mængde": "1.00",
                        "fakturabeløb": "25000.00",
                        "afgiftsbeløb": "5000.00",
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats/1":
            json_content = {
                "id": 1,
                "afgiftstabel": 1,
                "vareart_da": "Båthorn",
                "vareart_kl": "Båthorn",
                "afgiftsgruppenummer": 1234567,
                "enhed": "kg",
                "afgiftssats": "1.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            }
        elif path == expected_prefix + "notat":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "privatafgiftsanmeldelse": None,
                        "oprettet": "2023-10-01T00:00:00.000000+00:00",
                        "tekst": "Test tekst",
                        "index": 0,
                    }
                ],
            }
        elif path == expected_prefix + "prismeresponse":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "rec_id": "5637147578",
                        "tax_notification_number": "44668899",
                        "delivery_date": "2023-04-07T00:00:00",
                    }
                ],
            }
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    @patch.object(TemplateView, "get_context_data")
    def test_context_data(self, mock_super_context_data, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        try:
            self.client.get(reverse("tf10_history_view", kwargs={"id": 1, "index": 0}))
            mock_super_context_data.assert_called_with(
                {
                    "id": 1,
                    "index": 0,
                    "object": {
                        "id": 1,
                        "afsender": {
                            "id": 1,
                            "navn": "Testfirma1",
                            "adresse": "Testvej 1",
                            "postnummer": 1234,
                            "by": "Testby",
                            "postbox": None,
                            "telefon": "123456",
                            "cvr": 12345678,
                        },
                        "modtager": {
                            "id": 2,
                            "navn": "Testfirma2",
                            "adresse": "Testvej 2",
                            "postnummer": 1234,
                            "by": "Testby",
                            "postbox": None,
                            "telefon": "789012",
                            "cvr": 12345679,
                            "kreditordning": False,
                        },
                        "fragtforsendelse": {
                            "id": 1,
                            "forsendelsestype": "F",
                            "fragtbrevsnummer": "0",
                            "fragtbrev": "/fragtbreve/1/fragtbrev.txt",
                            "forbindelsesnr": "7331",
                            "afgangsdato": "2023-10-01",
                        },
                        "postforsendelse": None,
                        "leverandørfaktura_nummer": "1234",
                        "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                        "betales_af": "afsender",
                        "indførselstilladelse": "5678",
                        "afgift_total": "5000.00",
                        "betalt": False,
                        "dato": "2023-10-06T00:00:00-02:00",
                        "beregnet_faktureringsdato": "2023-10-10",
                        "status": "ny",
                        "history_username": "admin",
                        "history_date": "2023-10-01T00:00:00.000000+00:00",
                        "varelinjer": [
                            {
                                "id": 1,
                                "afgiftsanmeldelse": 1,
                                "vareafgiftssats": Vareafgiftssats(
                                    id=1,
                                    afgiftstabel=1,
                                    vareart_da="Båthorn",
                                    vareart_kl="Båthorn",
                                    afgiftsgruppenummer=1234567,
                                    enhed=Vareafgiftssats.Enhed.KILOGRAM,
                                    afgiftssats=Decimal("1.00"),
                                    kræver_indførselstilladelse=False,
                                    minimumsbeløb=None,
                                    overordnet=None,
                                    segment_nedre=None,
                                    segment_øvre=None,
                                    subsatser=None,
                                ),
                                "antal": 5,
                                "mængde": "1.00",
                                "fakturabeløb": "25000.00",
                                "afgiftsbeløb": "5000.00",
                            }
                        ],
                        "notater": [
                            Notat(
                                id=1,
                                tekst="Test tekst",
                                afgiftsanmeldelse=1,
                                oprettet=datetime(
                                    2023, 10, 1, 0, 0, tzinfo=timezone.utc
                                ),
                                navn=None,
                                index=0,
                            )
                        ],
                    },
                    "user": {
                        "id": 1,
                        "username": "admin",
                        "first_name": "Administrator",
                        "last_name": "",
                        "email": "admin@told.gl",
                        "is_superuser": True,
                        "groups": [],
                        "permissions": [
                            "auth.add_group",
                            "auth.change_group",
                            "auth.delete_group",
                            "auth.view_group",
                            "auth.add_permission",
                            "auth.change_permission",
                            "auth.delete_permission",
                            "auth.view_permission",
                            "auth.add_user",
                            "auth.change_user",
                            "auth.delete_user",
                            "auth.view_user",
                            "contenttypes.add_contenttype",
                            "contenttypes.change_contenttype",
                            "contenttypes.delete_contenttype",
                            "contenttypes.view_contenttype",
                            "sessions.add_session",
                            "sessions.change_session",
                            "sessions.delete_session",
                            "sessions.view_session",
                        ],
                    },
                }
            )
        except TypeError:
            pass  # Vi overstyrer TemplateView.get_context_data for at detektere og asserte på dens input,
            # men det betyder at den ikke rigtigt renderer siden. Det er vi dog. p.t. ligeglade med


class AnmeldelseNotatTest(PermissionsTest, TestCase):
    view = TF10FormUpdateView
    check_permissions = (
        (
            reverse("tf10_edit", kwargs={"id": 1}),
            view.required_permissions,
        ),
    )

    def setUp(self):
        super().setUp()
        self.patched: List[Tuple[str, str]] = []
        self.posted: List[Tuple[str, str]] = []

    @staticmethod
    def mock_requests_get(path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "vareafgiftssats":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftstabel": 1,
                        "vareart_da": "Båthorn",
                        "vareart_kl": "Båthorn",
                        "afgiftsgruppenummer": 1234567,
                        "enhed": "kg",
                        "afgiftssats": "1.00",
                        "kræver_indførselstilladelse": False,
                        "har_privat_tillægsafgift_alkohol": False,
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats/1":
            json_content = {
                "id": 1,
                "afgiftstabel": 1,
                "vareart_da": "Båthorn",
                "vareart_kl": "Båthorn",
                "afgiftsgruppenummer": 1234567,
                "enhed": "kg",
                "afgiftssats": "1.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            }

        elif path == expected_prefix + "afsender/1":
            json_content = {
                "id": 1,
                "navn": "Testfirma 1",
                "adresse": "Testvej 42",
                "postnummer": 1234,
                "by": "TestBy",
                "postbox": "123",
                "telefon": "123456",
                "cvr": 12345678,
            }
        elif path == expected_prefix + "afsender":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "navn": "Testfirma 1",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                    }
                ],
            }

        elif path == expected_prefix + "modtager/1":
            json_content = {
                "id": 1,
                "navn": "Testfirma 1",
                "adresse": "Testvej 42",
                "postnummer": 1234,
                "by": "TestBy",
                "postbox": "123",
                "telefon": "123456",
                "cvr": 12345678,
                "kreditordning": True,
            }

        elif path == expected_prefix + "modtager":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "navn": "Testfirma 1",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                        "kreditordning": True,
                    }
                ],
            }

        elif path == expected_prefix + "afgiftsanmeldelse/1":
            json_content = {
                "id": 1,
                "afsender": 1,
                "modtager": 1,
                "fragtforsendelse": 1,
                "postforsendelse": None,
                "leverandørfaktura_nummer": "1234",
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "betales_af": "afsender",
                "indførselstilladelse": "5678",
                "afgift_total": "5000.00",
                "betalt": False,
                "dato": "2023-08-22T00:00:00-02:00",
                "beregnet_faktureringsdato": "2023-10-10",
                "status": "afvist",
            }
        elif path == expected_prefix + "afgiftsanmeldelse/1/full":
            json_content = {
                "id": 1,
                "leverandørfaktura_nummer": "1234",
                "betales_af": "afsender",
                "indførselstilladelse": "abcde",
                "betalt": False,
                "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "afsender": {
                    "id": 20,
                    "navn": "Testfirma 5",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                },
                "modtager": {
                    "id": 21,
                    "navn": "Testfirma 3",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                    "kreditordning": True,
                },
                "postforsendelse": {
                    "id": 1,
                    "postforsendelsesnummer": "1234",
                    "forsendelsestype": "S",
                    "afsenderbykode": "2468",
                    "afgangsdato": "2023-11-03",
                },
                "dato": "2023-09-03T00:00:00-02:00",
                "beregnet_faktureringsdato": "2023-10-10",
                "afgift_total": "5000.00",
                "fragtforsendelse": None,
                "status": "afvist",
            }
        elif path == expected_prefix + "fragtforsendelse/1":
            json_content = {
                "id": 1,
                "forsendelsestype": "S",
                "fragtbrevsnummer": "ABCDE1234567",
                "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "forbindelsesnr": "ABC 337",
                "afgangsdato": "2023-10-01",
            }

        elif path == expected_prefix + "varelinje":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "vareafgiftssats": 1,
                        "antal": 5,
                        "mængde": 2,
                        "fakturabeløb": "25000.00",
                        "afgiftsbeløb": "5000.00",
                    }
                ],
            }

        elif path == expected_prefix + "notat":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "privatafgiftsanmeldelse": None,
                        "privatafgiftsanmeldelse": None,
                        "tekst": "Hephey",
                        "navn": "tester",
                        "oprettet": "2023-10-01T00:00:00.000000+00:00",
                        "index": 0,
                    }
                ],
            }
        elif path == expected_prefix + "afgiftstabel":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "gyldig_fra": "2022-01-01",
                        "gyldig_til": "2023-01-01",
                        "kladde": False,
                    },
                ],
            }
        elif path == expected_prefix + "toldkategori":
            json_content = [
                {
                    "kategori": "70",
                    "navn": "RAL Royal Arctic Line A/S",
                    "kræver_cvr": False,
                }
            ]
        else:
            raise Exception(f"Mock AnmeldelseNotatTest got unrecognized path: {path}")
            traceback.print_stack()
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_patch(self, path, data, headers=None):
        path = path.rstrip("/")
        response = Response()
        content = None
        status_code = None
        json_content = {"id": 1}
        self.patched.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_post(self, path, data, headers=None):
        path = path.rstrip("/")
        response = Response()
        content = None
        status_code = None
        json_content = {"id": 1}
        self.posted.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests.sessions.Session, "patch")
    @patch.object(requests.sessions.Session, "post")
    def test_add_notat(self, mock_post, mock_patch, mock_get):
        self.login()
        mock_get.side_effect = self.mock_requests_get
        mock_patch.side_effect = self.mock_requests_patch
        mock_post.side_effect = self.mock_requests_post
        with patch("builtins.open", mock_open(read_data=b"test_data")):
            response = self.client.post(
                reverse("tf10_edit", kwargs={"id": 1}),
                {
                    "afsender_cvr": "12345678",
                    "afsender_navn": "Testfirma 5",
                    "afsender_adresse": "Testvej 42",
                    "afsender_postbox": "123",
                    "afsender_postnummer": "1234",
                    "afsender_by": "TestBy",
                    "afsender_telefon": "123456",
                    "modtager_cvr": "12345678",
                    "modtager_navn": "Testfirma 3",
                    "modtager_adresse": "Testvej 42",
                    "modtager_postbox": "123",
                    "modtager_postnummer": ["3506"],
                    "modtager_by": "TestBy",
                    "modtager_telefon": "123456",
                    "indførselstilladelse": "123",
                    "leverandørfaktura_nummer": "5678",
                    "fragttype": "skibspost",
                    "forbindelsesnr": "1337",
                    "fragtbrevnr": "ABCDE1234567",
                    "afgangsdato": "2023-11-03",
                    "form-TOTAL_FORMS": "1",
                    "form-INITIAL_FORMS": "1",
                    "form-MIN_NUM_FORMS": "1",
                    "form-MAX_NUM_FORMS": "1000",
                    "form-0-id": 1,
                    "form-0-vareafgiftssats": 1,
                    "form-0-mængde": 2,
                    "form-0-antal": 5,
                    "form-0-fakturabeløb": "25000.00",
                    "notat": "Testnotat",
                },
            )
            self.assertIn(
                (
                    f"{settings.REST_DOMAIN}/api/notat",
                    '{"afgiftsanmeldelse_id": 1, "privatafgiftsanmeldelse_id": null, "tekst": "Testnotat"}',
                ),
                self.posted,
            )


class AdminFileViewTest(FileViewTest, TestCase):
    @property
    def login_url(self):
        return str(reverse("login"))

    @property
    def file_view_url(self):
        return str(reverse("fragtbrev_view", kwargs={"id": 1}))

    @property
    def file_view_url_2(self):
        return str(reverse("fragtbrev_view", kwargs={"id": 2}))


class AfgiftstabelListViewTest(PermissionsTest, TestCase):
    @property
    def login_url(self):
        return str(reverse("login"))

    view = AfgiftstabelListView
    check_permissions = ((reverse("afgiftstabel_list"), view.required_permissions),)

    @property
    def list_url(self):
        return reverse("afgiftstabel_list")

    def view_url(self, id: int):
        return reverse("afgiftstabel_view", kwargs={"id": id})

    @staticmethod
    def get_html_list(html: str):
        soup = BeautifulSoup(html, "html.parser")
        headers = [element.text for element in soup.css.select("table thead tr th")]
        return [
            dict(zip(headers, [td.text.strip() for td in row.select("td")]))
            for row in soup.css.select("table tbody tr")
        ]

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.testdata = [
            {
                "id": 1,
                "gyldig_fra": "2023-01-01T00:00:00-03:00",
                "gyldig_til": "2024-01-01T00:00:00-02:00",
                "kladde": False,
            },
            {
                "id": 2,
                "gyldig_fra": "2024-01-01T00:00:00-02:00",
                "gyldig_til": None,
                "kladde": False,
            },
            {
                "id": 3,
                "gyldig_fra": None,
                "gyldig_til": None,
                "kladde": True,
            },
        ]

    def mock_requests_get(self, path):
        expected_prefix = "/api/"
        p = urlparse(path)
        path = p.path
        query = parse_qs(p.query)
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel":
            items = deepcopy(self.testdata)
            if "offset" in query:
                items = items[int(query["offset"][0]) :]
            if "limit" in query:
                items = items[: int(query["limit"][0])]
            sort = query.get("sort")
            reverse = query.get("order") == ["desc"]
            if sort == ["gyldig_fra"]:
                items.sort(
                    key=lambda x: (x["gyldig_fra"] is None, x["gyldig_fra"], x["id"]),
                    reverse=reverse,
                )
            elif sort == ["gyldig_til"]:
                items.sort(
                    key=lambda x: (x["gyldig_til"] is None, x["gyldig_til"], x["id"]),
                    reverse=reverse,
                )
            elif sort == ["kladde"]:
                items.sort(key=lambda x: (x["kladde"], x["id"]), reverse=reverse)

            json_content = {"count": len(items), "items": items}
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def test_requires_login(self):
        url = self.list_url
        response = self.client.get(url)
        self.assertEquals(response.status_code, 302)
        self.assertEquals(
            response.headers["Location"],
            self.login_url + "?back=" + quote(url, safe=""),
        )

    @patch.object(requests.sessions.Session, "get")
    def test_list(self, mock_get):
        self.maxDiff = None
        mock_get.side_effect = self.mock_requests_get
        self.login()
        url = self.list_url
        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)
        table_data = self.get_html_list(response.content)
        self.assertEquals(
            # Tag table_data og fjern alle html-tags i strengen. Join med mellemrum.
            self.strip_html_tags(table_data),
            [
                {
                    "Gyldig fra": "2023-01-01T00:00:00-03:00",
                    "Gyldig til": "2024-01-01T00:00:00-02:00",
                    "Kladde": "Nej",
                    "Handlinger": "Vis Download .xlsx .csv",
                },
                {
                    "Gyldig fra": "2024-01-01T00:00:00-02:00",
                    "Gyldig til": "-",
                    "Kladde": "Nej",
                    "Handlinger": "Vis Download .xlsx .csv",
                },
                {
                    "Gyldig fra": "-",
                    "Gyldig til": "-",
                    "Kladde": "Ja",
                    "Handlinger": "Vis Download .xlsx .csv",
                },
            ],
        )

        url = self.list_url + "?json=1"
        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)
        data = response.json()

        self.assertEquals(
            self.strip_html_tags(data),
            {
                "total": 3,
                "items": [
                    {
                        "id": 1,
                        "gyldig_fra": "2023-01-01T00:00:00-03:00",
                        "gyldig_til": "2024-01-01T00:00:00-02:00",
                        "kladde": False,
                        "actions": "Vis Download .xlsx .csv",
                        "gældende": False,
                    },
                    {
                        "id": 2,
                        "gyldig_fra": "2024-01-01T00:00:00-02:00",
                        "gyldig_til": None,
                        "kladde": False,
                        "actions": "Vis Download .xlsx .csv",
                        "gældende": True,
                    },
                    {
                        "id": 3,
                        "gyldig_fra": None,
                        "gyldig_til": None,
                        "kladde": True,
                        "actions": "Vis Download .xlsx .csv",
                        "gældende": False,
                    },
                ],
            },
        )

    @staticmethod
    def strip_html_tags(data: Dict) -> Dict:
        # Tag table_data og fjern alle html-tags i alle streng-values. Join med mellemrum.
        return modify_values(
            data,
            (str,),
            lambda s: " ".join(
                filter(
                    None,
                    [x.strip() for x in re.sub("<[^>]+>", "", s).split("\n")],
                )
            ),
        )

    @patch.object(requests.sessions.Session, "get")
    def test_list_sort(self, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        sort_tests = [
            ("gyldig_fra", "", [1, 2, 3]),
            ("gyldig_fra", "asc", [1, 2, 3]),
            ("gyldig_fra", "desc", [3, 2, 1]),
            ("gyldig_til", "", [1, 2, 3]),
            ("gyldig_til", "asc", [1, 2, 3]),
            ("gyldig_til", "desc", [3, 2, 1]),
            ("kladde", "", [1, 2, 3]),
            ("kladde", "asc", [1, 2, 3]),
            ("kladde", "desc", [3, 2, 1]),
        ]
        for test in sort_tests:
            url = self.list_url + f"?json=1&sort={test[0]}&order={test[1]}"
            response = self.client.get(url)
            numbers = [int(item["id"]) for item in response.json()["items"]]
            self.assertEquals(response.status_code, 200)
            self.assertEquals(numbers, test[2], f"{test[0]} {test[1]}")

    @patch.object(requests.sessions.Session, "get")
    def test_list_paginate(self, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        paginate_tests = [
            (-1, 3, [1, 2, 3]),
            (0, 3, [1, 2, 3]),
            (1, 3, [2, 3]),
            (2, 3, [3]),
            (0, 2, [1, 2]),
            (1, 2, [2, 3]),
            (2, 2, [3]),
            (0, 1, [1]),
            (1, 1, [2]),
            (2, 1, [3]),
            (0, 0, [1]),
            (1, 0, [2]),
            (2, 0, [3]),
        ]
        for offset, limit, expected in paginate_tests:
            url = self.list_url + f"?json=1&offset={offset}&limit={limit}"
            response = self.client.get(url)
            numbers = [int(item["id"]) for item in response.json()["items"]]
            self.assertEquals(response.status_code, 200)
            self.assertEquals(numbers, expected)


class AfgiftstabelDetailViewTest(PermissionsTest, TestCase):
    view = AfgiftstabelDetailView
    check_permissions = (
        (reverse("afgiftstabel_view", kwargs={"id": 1}), view.required_permissions),
    )

    def setUp(self):
        super().setUp()
        self.patched: List[Tuple[str, str]] = []

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.afgiftstabel1 = {
            "id": 1,
            "gyldig_fra": "2022-01-01T00:00:00-03:00",
            "gyldig_til": None,
            "kladde": True,
        }
        cls.afgiftstabel2 = {
            "id": 1,
            "gyldig_fra": "2022-01-01T00:00:00-03:00",
            "gyldig_til": None,
            "kladde": False,
        }
        cls.afgiftssatser = [
            {
                "id": 1,
                "afgiftstabel": 1,
                "afgiftsgruppenummer": 70,
                "vareart_da": "FYRVÆRKERI",
                "vareart_kl": "FYRVÆRKERI",
                "enhed": "pct",
                "minimumsbeløb": False,
                "afgiftssats": "100.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            },
        ]

    def mock_requests_get(self, path):
        expected_prefix = "/api/"
        p = urlparse(path)
        path = p.path
        query = parse_qs(p.query)
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel/1":
            json_content = self.afgiftstabel1
        if path == expected_prefix + "afgiftstabel/2":
            json_content = self.afgiftstabel2
        if path == expected_prefix + "vareafgiftssats":
            json_content = {
                "count": len(self.afgiftssatser),
                "items": self.afgiftssatser,
            }
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_patch(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel/1":
            json_content = {"id": 1}
            self.patched.append((path, data))
        else:
            raise Exception(
                f"Mock {self.__class__.__name__} got unrecognized path: PATCH {path}"
            )
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def parse_form_row(self, soup, class_):
        row = soup.find("form").find("div", class_=class_)
        items = []
        for cell in row.children:
            strings = list(cell.stripped_strings)
            if isinstance(cell, Tag):
                for element in cell.children:
                    if isinstance(element, Tag):
                        if element.get("value"):
                            strings.append(element.get("value"))
            for x in strings:
                subitems = []
                if x:
                    subitems.append(x)
                if subitems:
                    items.append(subitems)
        return items

    @patch.object(requests.sessions.Session, "get")
    def test_datetime_show(self, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        response = self.client.get(reverse("afgiftstabel_view", kwargs={"id": 2}))
        soup = BeautifulSoup(response.content, "html.parser")
        self.assertEquals(
            self.parse_form_row(soup, "gyldig_fra"),
            [["Gyldig fra"], ["2022-01-01T00:00:00-03:00"]],
        )
        self.assertEquals(
            self.parse_form_row(soup, "gyldig_til"), [["Gyldig til"], ["-"]]
        )

    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests.sessions.Session, "patch")
    def test_datetime_update(self, mock_patch, mock_get):
        mock_get.side_effect = self.mock_requests_get
        mock_patch.side_effect = self.mock_requests_patch
        self.login()
        self.client.post(
            reverse("afgiftstabel_view", kwargs={"id": 1}),
            {
                "gyldig_fra": "11/01/2025 00:00",
                "offset": "60",
                "kladde": True,
            },
        )
        prefix = f"{settings.REST_DOMAIN}/api/"
        patched_map = defaultdict(list)
        for url, data in self.patched:
            patched_map[url].append(json.loads(data))
        self.assertEquals(
            patched_map[prefix + "afgiftstabel/1"],
            [
                {
                    "kladde": "True",
                    "gyldig_fra": "2025-01-11T00:00:00+01:00",
                    "offset": 60,
                    "delete": False,
                }
            ],
        )


class AfgiftstabelDownloadTest(PermissionsTest, TestCase):
    @property
    def login_url(self):
        return str(reverse("login"))

    view = AfgiftstabelDownloadView
    check_permissions = (
        (
            reverse("afgiftstabel_download", kwargs={"id": 1, "format": "xlsx"}),
            view.required_permissions,
        ),
        (
            reverse("afgiftstabel_download", kwargs={"id": 1, "format": "csv"}),
            view.required_permissions,
        ),
    )

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.afgiftstabel = {
            "id": 1,
            "gyldig_fra": "2022-01-01T00:00:00-02:00",
            "gyldig_til": "2023-01-01T00:00:00-02:00",
            "kladde": False,
        }
        cls.afgiftssatser = [
            {
                "id": 1,
                "afgiftstabel": 1,
                "afgiftsgruppenummer": 70,
                "vareart_da": "FYRVÆRKERI",
                "vareart_kl": "FYRVÆRKERI",
                "enhed": "pct",
                "minimumsbeløb": False,
                "afgiftssats": "100.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            },
            {
                "id": 2,
                "afgiftstabel": 1,
                "afgiftsgruppenummer": 71,
                "vareart_da": "KNALLERTER",
                "vareart_kl": "KNALLERTER",
                "enhed": "ant",
                "minimumsbeløb": False,
                "afgiftssats": "2530.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            },
        ]

    def mock_requests_get(self, path):
        expected_prefix = "/api/"
        p = urlparse(path)
        path = p.path
        query = parse_qs(p.query)
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel/1":
            json_content = self.afgiftstabel
        if path == expected_prefix + "vareafgiftssats":
            json_content = {
                "count": len(self.afgiftssatser),
                "items": self.afgiftssatser,
            }
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    def test_xlsx(self, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        response = self.client.get(
            reverse("afgiftstabel_download", kwargs={"id": 1, "format": "xlsx"})
        )
        self.assertEquals(response.status_code, 200)
        self.assertEquals(
            response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(len(response.content) > 0)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEquals(
            list(workbook.active.values),
            [
                (
                    "Afgiftsgruppenummer",
                    "Overordnet",
                    "Vareart (da)",
                    "Vareart (kl)",
                    "Enhed",
                    "Afgiftssats",
                    "Kræver indførselstilladelse",
                    "Har privat tillægsafgift alkohol",
                    "Synlig for private",
                    "Minimumsbeløb",
                    "Segment nedre",
                    "Segment øvre",
                ),
                (
                    70,
                    None,
                    "FYRVÆRKERI",
                    "FYRVÆRKERI",
                    "procent",
                    "100,00",
                    "nej",
                    "nej",
                    "nej",
                    "0,00",
                    None,
                    None,
                ),
                (
                    71,
                    None,
                    "KNALLERTER",
                    "KNALLERTER",
                    "antal",
                    "2.530,00",
                    "nej",
                    "nej",
                    "nej",
                    "0,00",
                    None,
                    None,
                ),
            ],
        )

    @patch.object(requests.sessions.Session, "get")
    def test_csv(self, mock_get):
        mock_get.side_effect = self.mock_requests_get
        self.login()
        response = self.client.get(
            reverse("afgiftstabel_download", kwargs={"id": 1, "format": "csv"})
        )
        self.assertEquals(response.status_code, 200)
        self.assertEquals(response.headers["Content-Type"], "text/csv")
        self.assertTrue(len(response.content) > 0)
        reader = csv.reader(StringIO(response.content.decode("utf-8")))
        self.assertEquals(
            [row for row in reader],
            [
                [
                    "Afgiftsgruppenummer",
                    "Overordnet",
                    "Vareart (da)",
                    "Vareart (kl)",
                    "Enhed",
                    "Afgiftssats",
                    "Kræver indførselstilladelse",
                    "Har privat tillægsafgift alkohol",
                    "Synlig for private",
                    "Minimumsbeløb",
                    "Segment nedre",
                    "Segment øvre",
                ],
                [
                    "70",
                    "",
                    "FYRVÆRKERI",
                    "FYRVÆRKERI",
                    "procent",
                    "100,00",
                    "nej",
                    "nej",
                    "nej",
                    "0,00",
                    "",
                    "",
                ],
                [
                    "71",
                    "",
                    "KNALLERTER",
                    "KNALLERTER",
                    "antal",
                    "2.530,00",
                    "nej",
                    "nej",
                    "nej",
                    "0,00",
                    "",
                    "",
                ],
            ],
        )


class AfgiftstabelUploadTest(TestMixin, HasLogin):
    @property
    def login_url(self):
        return str(reverse("login"))

    def setUp(self):
        super().setUp()
        self.posted = []
        self.id_counter = 1

        self.data = [
            [
                "Afgiftsgruppenummer",
                "Overordnet",
                "Vareart (da)",
                "Vareart (kl)",
                "Enhed",
                "Afgiftssats",
                "Kræver indførselstilladelse",
                "Har privat tillægsafgift alkohol",
                "Synlig for private",
                "Minimumsbeløb",
                "Segment nedre",
                "Segment øvre",
            ],
            [
                "1",
                "",
                "SUKKER og sirup",
                "SUKKER og sirup",
                "kilogram",
                "6,00",
                "nej",
                "nej",
                "nej",
                "0,00",
                "",
                "",
            ],
            [
                "2",
                "",
                "KAFFE, pulverkaffe, koncentrater",
                "KAFFE, pulverkaffe, koncentrater",
                "kilogram",
                "6,00",
                "nej",
                "nej",
                "nej",
                "0,00",
                "",
                "",
            ],
            [
                "3",
                "",
                "THE, pulver The, koncentrater",
                "THE, pulver The, koncentrater",
                "kilogram",
                "6,60",
                "nej",
                "nej",
                "nej",
                "0,00",
                "",
                "",
            ],
        ]

    def mock_requests_post(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel":
            json_content = {"id": 1}
            self.posted.append((path, data))
        if path == expected_prefix + "vareafgiftssats":
            json_content = {"id": self.id_counter}
            self.id_counter += 1
            self.posted.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "post")
    def test_successful_upload(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        response = self.upload(self.data)
        self.assertEquals(response.status_code, 302)
        self.assertEquals(response.headers["Location"], reverse("afgiftstabel_list"))

    @patch.object(requests.sessions.Session, "post")
    def test_failure_wrong_content_type(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        response = self.upload(self.data, "text/plain")
        self.assertEquals(response.status_code, 200)
        errors = self.get_errors(response.content)
        self.assertTrue("fil" in errors)
        self.assertEquals(errors["fil"], ["Ugyldig content-type: text/plain"])

    @patch.object(requests.sessions.Session, "post")
    def test_failure_header_missing(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        for i in range(0, 9):
            data = deepcopy(self.data)
            for line in data:
                line.pop(i)
            response = self.upload(data)
            self.assertEquals(response.status_code, 200)
            errors = self.get_errors(response.content)
            self.assertTrue("fil" in errors)
            self.assertEquals(errors["fil"], [f"Mangler kolonne med {self.data[0][i]}"])

    @patch.object(requests.sessions.Session, "post")
    def test_failure_number_twice(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        self.data[3][0] = "2"
        response = self.upload(self.data)
        self.assertEquals(response.status_code, 200)
        errors = self.get_errors(response.content)
        self.assertTrue("fil" in errors)
        self.assertEquals(
            errors["fil"], ["Afgiftsgruppenummer 2 optræder to gange (linjer: 3, 4)"]
        )

    @patch.object(requests.sessions.Session, "post")
    def test_failure_missing_reference(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        self.data[3][1] = "4"
        response = self.upload(self.data)
        self.assertEquals(response.status_code, 200)
        errors = self.get_errors(response.content)
        self.assertTrue("fil" in errors)
        self.assertEquals(
            errors["fil"],
            [
                "Afgiftssats med afgiftsgruppenummer 3 (linje 4) peger på overordnet 4, som ikke findes"
            ],
        )

    @patch.object(requests.sessions.Session, "post")
    def test_failure_reference_self(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        self.data[1][1] = "1"
        response = self.upload(self.data)
        self.assertEquals(response.status_code, 200)
        errors = self.get_errors(response.content)
        self.assertTrue("fil" in errors)
        self.assertEquals(
            errors["fil"],
            ["Vareafgiftssats 1 (linje 2) peger på sig selv som overordnet"],
        )

    @patch.object(requests.sessions.Session, "post")
    def test_failure_circular_reference(self, mock_post):
        mock_post.side_effect = self.mock_requests_post
        self.data[1][1] = "3"
        self.data[2][1] = "1"
        self.data[3][1] = "2"
        response = self.upload(self.data)
        self.assertEquals(response.status_code, 200)
        errors = self.get_errors(response.content)
        self.assertTrue("fil" in errors)
        self.assertEquals(
            errors["fil"],
            [
                "Vareafgiftssats 2 (linje 3) har 1 (linje 2) som overordnet, men 1 har også 2 i kæden af overordnede"
            ],
        )


class AfgiftstabelUploadCsvTest(AfgiftstabelUploadTest, TestCase):
    def upload(self, data: List[List[str]], content_type: str = "text/csv"):
        encoded = "\n".join(
            [",".join([f'"{cell}"' for cell in line]) for line in data]
        ).encode("utf-8")
        self.login()
        return self.client.post(
            reverse("afgiftstabel_create"),
            {"fil": SimpleUploadedFile("test.csv", encoded, content_type)},
        )


class AfgiftstabelUploadXlsxTest(AfgiftstabelUploadTest, TestCase):
    def upload(
        self,
        data: List[List[str]],
        content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ):
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        for item in data:
            ws.append(item)
        output = BytesIO()
        wb.save(output)
        encoded = output.getvalue()
        self.login()
        return self.client.post(
            reverse("afgiftstabel_create"),
            {"fil": SimpleUploadedFile("test.xlsx", encoded, content_type)},
        )


class TF10EditMultipleViewTest(PermissionsTest, TestCase):
    view = TF10EditMultipleView
    check_permissions = (
        (reverse("tf10_edit_multiple") + "?id=1", view.required_permissions, 302),
        (reverse("tf10_edit_multiple") + "?id=1&id=2", view.required_permissions),
    )

    def mock_requests_get(self, path):
        expected_prefix = "/api/"
        p = urlparse(path)
        path = p.path
        query = parse_qs(p.query)
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftsanmeldelse/full":
            data = {
                1: {
                    "id": 1,
                    "afsender": {
                        "id": 20,
                        "navn": "Testfirma 5",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                    },
                    "modtager": {
                        "id": 21,
                        "navn": "Testfirma 3",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                        "kreditordning": True,
                    },
                    "fragtforsendelse": {
                        "id": 1,
                        "forsendelsestype": "S",
                        "fragtbrevsnummer": "ABCDE1234567",
                        "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                        "forbindelsesnr": "ABC 337",
                        "afgangsdato": "2023-10-01",
                    },
                    "postforsendelse": None,
                    "leverandørfaktura_nummer": "1234",
                    "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                    "betales_af": "afsender",
                    "indførselstilladelse": "5678",
                    "afgift_total": "5000.00",
                    "betalt": False,
                    "dato": "2023-08-22T00:00:00-02:00",
                    "beregnet_faktureringsdato": "2023-10-10",
                    "status": "afvist",
                },
                2: {
                    "id": 2,
                    "afsender": {
                        "id": 20,
                        "navn": "Testfirma 5",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                    },
                    "modtager": {
                        "id": 21,
                        "navn": "Testfirma 3",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                        "kreditordning": True,
                    },
                    "fragtforsendelse": {
                        "id": 2,
                        "forsendelsestype": "F",
                        "fragtbrevsnummer": 1,
                        "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                        "forbindelsesnr": "123",
                        "afgangsdato": "2023-10-01",
                    },
                    "postforsendelse": None,
                    "leverandørfaktura_nummer": "1234",
                    "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                    "betales_af": "afsender",
                    "indførselstilladelse": "5678",
                    "afgift_total": "5000.00",
                    "betalt": False,
                    "dato": "2023-08-22T00:00:00-02:00",
                    "beregnet_faktureringsdato": "2023-10-10",
                    "status": "afvist",
                },
                3: {
                    "id": 3,
                    "afsender": {
                        "id": 20,
                        "navn": "Testfirma 5",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                    },
                    "modtager": {
                        "id": 21,
                        "navn": "Testfirma 3",
                        "adresse": "Testvej 42",
                        "postnummer": 1234,
                        "by": "TestBy",
                        "postbox": "123",
                        "telefon": "123456",
                        "cvr": 12345678,
                        "kreditordning": True,
                    },
                    "fragtforsendelse": {
                        "id": 3,
                        "forsendelsestype": "F",
                        "fragtbrevsnummer": 1,
                        "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                        "forbindelsesnr": "123",
                        "afgangsdato": "2023-10-01",
                    },
                    "postforsendelse": None,
                    "leverandørfaktura_nummer": "1234",
                    "leverandørfaktura": "/leverandørfakturaer/1/leverandørfaktura.txt",
                    "betales_af": "afsender",
                    "indførselstilladelse": "5678",
                    "afgift_total": "5000.00",
                    "betalt": False,
                    "dato": "2023-08-22T00:00:00-02:00",
                    "beregnet_faktureringsdato": "2023-10-10",
                    "status": "afvist",
                },
            }
            items = list(filter(None, [data.get(int(id)) for id in query["id"]]))
            json_content = {
                "count": len(items),
                "items": items,
            }
        elif path == expected_prefix + "fragtforsendelse/1":
            json_content = {
                "id": 1,
                "forsendelsestype": "S",
                "fragtbrevsnummer": "ABCDE1234567",
                "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "forbindelsesnr": "ABC 337",
                "afgangsdato": "2023-10-01",
            }
        elif path == expected_prefix + "fragtforsendelse/2":
            json_content = {
                "id": 2,
                "forsendelsestype": "F",
                "fragtbrevsnummer": 1,
                "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "forbindelsesnr": "123",
                "afgangsdato": "2023-10-01",
            }
        elif path == expected_prefix + "fragtforsendelse/3":
            json_content = {
                "id": 3,
                "forsendelsestype": "F",
                "fragtbrevsnummer": 1,
                "fragtbrev": "/leverandørfakturaer/1/leverandørfaktura.txt",
                "forbindelsesnr": "123",
                "afgangsdato": "2023-10-01",
            }
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: {path}")

        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def analyze_html(self, html):
        soup = BeautifulSoup(html, "html.parser")
        return {
            "fields": [x.attrs["name"] for x in soup.find_all("input")],
            "disabled_fields": [
                x.attrs["name"] for x in soup.find_all("input", disabled=True)
            ],
            "alerts": [x.text.strip() for x in soup.find_all(class_="alert")],
        }

    def test_get_redirect(self):
        self.login()
        response = self.client.get(reverse("tf10_edit_multiple") + "?id=1")
        self.assertEquals(response.status_code, 302)
        self.assertEquals(
            response.headers["Location"], reverse("tf10_edit", kwargs={"id": 1})
        )

    @patch.object(requests.sessions.Session, "get")
    def test_get_form_diff_fragttype(self, mock_get):
        self.login()
        mock_get.side_effect = self.mock_requests_get
        response = self.client.get(reverse("tf10_edit_multiple") + "?id=1&id=2")
        self.assertEquals(response.status_code, 200)
        analysis = self.analyze_html(response.content)
        self.assertIn("forbindelsesnr", analysis["disabled_fields"])
        self.assertIn("afgangsdato", analysis["disabled_fields"])
        self.assertIn(
            "Forbindelsesnummer, Afsenderbykode, Fragtbrevnr, Postforsendelsesnummer "
            "og Afgangsdato kan kun redigeres hvis alle de redigerede afgiftsanmeldelser "
            "har samme fragttype.",
            analysis["alerts"],
        )

    @patch.object(requests.sessions.Session, "get")
    def test_get_form_same_fragttype(self, mock_get):
        self.login()
        mock_get.side_effect = self.mock_requests_get
        response = self.client.get(reverse("tf10_edit_multiple") + "?id=2&id=3")
        self.assertEquals(response.status_code, 200)
        analysis = self.analyze_html(response.content)
        self.assertNotIn("forbindelsesnr", analysis["disabled_fields"])
        self.assertNotIn("fragtbrevnr", analysis["disabled_fields"])
        self.assertNotIn("afgangsdato", analysis["disabled_fields"])
        self.assertNotIn(
            "Forbindelsesnummer, Afsenderbykode, Fragtbrevnr, Postforsendelsesnummer "
            "og Afgangsdato kan kun redigeres hvis alle de redigerede afgiftsanmeldelser "
            "har samme fragttype.",
            analysis["alerts"],
        )


class TF10CreateTest(TestMixin, HasLogin, TestCase):
    @property
    def login_url(self):
        return str(reverse("login:login"))

    formdata1 = {
        "afsender_cvr": "12345678",
        "afsender_navn": "TestFirma1",
        "afsender_adresse": "Testvej 42",
        "afsender_postnummer": "1234",
        "afsender_by": "TestBy",
        "afsender_postbox": "123",
        "afsender_telefon": "123456",
        "modtager_cvr": "12345679",
        "modtager_navn": "TestFirma2",
        "modtager_adresse": "Testvej 43",
        "modtager_postnummer": "1234",
        "modtager_by": "TestBy",
        "modtager_postbox": "124",
        "modtager_telefon": "123123",
        "indførselstilladelse": "123",
        "leverandørfaktura_nummer": "123",
        "fragttype": "skibsfragt",
        "fragtbrevnr": "ABCDE1234567",
        "afgangsdato": "2023-11-03",
        "form-TOTAL_FORMS": "1",
        "form-INITIAL_FORMS": "1",
        "form-0-vareafgiftssats": "1",
        "form-0-mængde": "3",
        "form-0-antal": "6",
        "form-0-fakturabeløb": "100.00",
        "forbindelsesnr": "ABC 337",
        "oprettet_på_vegne_af": 1,
        "betales_af": "afsender",
    }
    formdata2 = {**formdata1, "fragttype": "luftpost", "forbindelsesnr": "1337"}

    _formfiles1 = {
        "fragtbrev": SimpleUploadedFile("fragtbrev.txt", "Testtekst".encode("utf-8")),
        "leverandørfaktura": SimpleUploadedFile(
            "leverandørfaktura.txt", "Testtekst".encode("utf-8")
        ),
    }
    _formfiles2 = {
        "leverandørfaktura": SimpleUploadedFile(
            "leverandørfaktura.txt", "Testtekst".encode("utf-8")
        ),
    }
    subformdata1 = {
        "vareafgiftssats": "1",
        "mængde": "3",
        "antal": "6",
        "fakturabeløb": "100,00",
    }

    @property
    def formfiles1(self):
        for file in self._formfiles1.values():
            file.seek(0)
        return self._formfiles1

    @property
    def formfiles2(self):
        for file in self._formfiles2.values():
            file.seek(0)
        return self._formfiles2

    def setUp(self):
        super().setUp()
        self.posted: List[Tuple[str, str]] = []
        self.mock_existing = {
            "afsender": False,
            "modtager": False,
            "vareafgiftssats": True,
            "user": True,
        }

    # def test_uploadfile_to_base64str(self):
    #     file = TemporaryUploadedFile("testfile", "text/plain", 8, "utf-8")
    #     with file.open("+") as fp:
    #         fp.write(b"testdata")
    #         file.seek(0)
    #         self.assertEquals(RestClient._uploadfile_to_base64str(file), "dGVzdGRhdGE=")
    #     file = SimpleUploadedFile("testfile", b"testdata", "text/plain")
    #     file.seek(0)
    #     self.assertEquals(RestClient._uploadfile_to_base64str(file), "dGVzdGRhdGE=")

    def test_requires_login(self):
        url = str(reverse("tf10_create"))
        response = self.client.get(url)
        self.assertEquals(response.status_code, 302)
        self.assertEquals(
            response.headers["Location"],
            reverse("login") + "?back=" + quote(url, safe=""),
        )

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        empty = {"count": 0, "items": []}
        if path == expected_prefix + "afgiftstabel":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "gyldig_fra": "2023-01-01T00:00:00-02:00",
                        "gyldig_til": None,
                        "kladde": False,
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats":
            if self.mock_existing["vareafgiftssats"]:
                json_content = {
                    "count": 1,
                    "items": [
                        {
                            "id": 1,
                            "afgiftstabel": 1,
                            "vareart_da": "Båthorn",
                            "vareart_kl": "Båthorn",
                            "afgiftsgruppenummer": 1234567,
                            "enhed": "kg",
                            "afgiftssats": "1.00",
                        }
                    ],
                }
            else:
                json_content = empty
        elif path == expected_prefix + "afsender":
            if self.mock_existing["afsender"]:
                json_content = {
                    "count": 1,
                    "items": [
                        {
                            "id": 1,
                            "navn": "Testfirma 1",
                            "adresse": "Testvej 42",
                            "postnummer": 1234,
                            "by": "TestBy",
                            "postbox": "123",
                            "telefon": "123456",
                            "cvr": 12345678,
                        }
                    ],
                }
            else:
                json_content = empty
        elif path == expected_prefix + "modtager":
            if self.mock_existing["modtager"]:
                json_content = {
                    "count": 1,
                    "items": [
                        {
                            "id": 1,
                            "navn": "Testfirma 1",
                            "adresse": "Testvej 42",
                            "postnummer": 1234,
                            "by": "TestBy",
                            "postbox": "123",
                            "telefon": "123456",
                            "cvr": 12345678,
                            "kreditordning": True,
                        }
                    ],
                }
            else:
                json_content = empty
        elif path == expected_prefix + "user":
            if self.mock_existing["user"]:
                json_content = {
                    "count": 1,
                    "items": [
                        {
                            "id": 1,
                            "username": "indberetter",
                            "first_name": "Testperson",
                            "last_name": "Testersen",
                            "email": "",
                            "is_superuser": False,
                            "groups": ["ErhvervIndberettere"],
                            "permissions": [
                                "aktør.add_afsender",
                                "aktør.add_modtager",
                                "aktør.view_afsender",
                                "aktør.view_modtager",
                                "anmeldelse.add_afgiftsanmeldelse",
                                "anmeldelse.add_varelinje",
                                "anmeldelse.change_afgiftsanmeldelse",
                                "anmeldelse.view_afgiftsanmeldelse",
                                "anmeldelse.view_all_anmeldelse",
                                "anmeldelse.view_varelinje",
                                "forsendelse.add_fragtforsendelse",
                                "forsendelse.add_postforsendelse",
                                "forsendelse.view_all_fragtforsendelser",
                                "forsendelse.view_all_postforsendelser",
                                "forsendelse.view_fragtforsendelse",
                                "forsendelse.view_postforsendelse",
                                "sats.view_afgiftstabel",
                                "sats.view_vareafgiftssats",
                            ],
                            "indberetter_data": {
                                "cpr": "1234567890",
                                "cvr": "12345678",
                            },
                            "twofactor_enabled": False,
                        }
                    ],
                }
            else:
                json_content = empty
        else:
            print(f"Mock {self.__class__.__name__} got unrecognized path: GET {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_post(self, path, data, headers=None):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        self.posted.append((path, data))
        if path in (
            expected_prefix + x
            for x in (
                "afsender",
                "modtager",
                "fragtforsendelse",
                "postforsendelse",
                "afgiftsanmeldelse",
                "varelinje",
            )
        ):
            json_content = {"id": 1}
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def submit_get_errors(self, url, data):
        return self.get_errors(self.client.post(url, data=data).content)

    @patch.object(requests.Session, "get")
    def test_get_form(self, mock_get):
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get

        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)
        soup = BeautifulSoup(response.content, "html.parser")
        input_field_names = set(
            [
                input_field["name"]
                for input_field in soup.find_all(name=["input", "select"])
                if input_field.has_attr("name")
            ]
        )
        for field_name in (
            "csrfmiddlewaretoken",
            "afsender_cvr",
            "afsender_navn",
            "afsender_adresse",
            "afsender_postnummer",
            "afsender_by",
            "afsender_postbox",
            "afsender_telefon",
            "modtager_cvr",
            "modtager_navn",
            "modtager_adresse",
            "modtager_postnummer",
            "modtager_by",
            "modtager_postbox",
            "modtager_telefon",
            "indførselstilladelse",
            "fragttype",
            "forbindelsesnr",
            "fragtbrevnr",
            "leverandørfaktura",
            "leverandørfaktura_nummer",
            "fragtbrev",
            "afgangsdato",
            "form-0-vareafgiftssats",
            "form-0-antal",
            "form-0-mængde",
            "form-0-fakturabeløb",
            "form-INITIAL_FORMS",
            "form-TOTAL_FORMS",
            "form-MIN_NUM_FORMS",
            "form-MAX_NUM_FORMS",
        ):
            self.assertIn(field_name, input_field_names, f"Missing {field_name}")

    @patch.object(requests.Session, "get")
    @patch("admin.views.TF10FormCreateView.form_valid")
    def test_form_required_fields(self, mock_form_valid, mock_get):
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get
        mock_form_valid.return_value = redirect(reverse("tf10_create"))
        for required_field in (
            "afsender_navn",
            "modtager_navn",
        ):
            data = {**self.formdata1}
            del data[required_field]
            form = TF10CreateForm(
                data=data, files=self.formfiles1, oprettet_på_vegne_af_choices=()
            )
            self.assertTrue(required_field in form.errors)
            self.assertEquals(form.errors[required_field], ["Dette felt er påkrævet."])
            html_errors = self.submit_get_errors(url, data)
            self.assertTrue(required_field in html_errors)
            self.assertEquals(html_errors[required_field], ["Dette felt er påkrævet."])

        files = {**self.formfiles1}
        del files["leverandørfaktura"]
        form = TF10CreateForm(
            data=self.formdata1, files=files, oprettet_på_vegne_af_choices=()
        )
        self.assertEquals(form.errors["leverandørfaktura"], ["Dette felt er påkrævet."])
        html_errors = self.submit_get_errors(url, {**self.formdata1, **files})
        self.assertEquals(html_errors["leverandørfaktura"], ["Dette felt er påkrævet."])

    def test_vareform_required_fields(self):
        varesatser = {
            1: Vareafgiftssats(
                id=1,
                afgiftstabel=1,
                vareart_da="Båthorn",
                vareart_kl="Båthorn",
                afgiftsgruppenummer=12345678,
                enhed=Vareafgiftssats.Enhed.KILOGRAM,
                afgiftssats="1.00",
            ),
            2: Vareafgiftssats(
                id=2,
                afgiftstabel=1,
                vareart_da="Klovnesko",
                vareart_kl="Klovnesko",
                afgiftsgruppenummer=87654321,
                enhed=Vareafgiftssats.Enhed.ANTAL,
                afgiftssats="1.00",
            ),
            3: Vareafgiftssats(
                id=3,
                afgiftstabel=1,
                vareart_da="Ethjulede cykler",
                vareart_kl="Ethjulede cykler",
                afgiftsgruppenummer=22446688,
                enhed=Vareafgiftssats.Enhed.PROCENT,
                afgiftssats="0.50",
            ),
        }

        for required_field in ("vareafgiftssats",):
            data = {**self.subformdata1}
            del data[required_field]
            form = TF10VareForm(data=data, varesatser=varesatser)
            self.assertTrue(required_field in form.errors)
            self.assertEquals(form.errors[required_field], ["Dette felt er påkrævet."])

        for required_field, vareart in (
            ("mængde", 1),
            ("antal", 2),
            ("fakturabeløb", 3),
        ):
            data = {**self.subformdata1, "vareafgiftssats": vareart}
            del data[required_field]
            form = TF10VareForm(data=data, varesatser=varesatser)
            self.assertTrue(required_field in form.errors, required_field)
            self.assertEquals(
                form.errors[required_field],
                ["Dette felt er påkrævet."],
                required_field,
            )

    @patch.object(requests.Session, "get")
    @patch.object(requests.Session, "post")
    def test_form_successful(self, mock_post, mock_get):
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get
        mock_post.side_effect = self.mock_requests_post
        response = self.client.post(url, data={**self.formdata1, **self.formfiles1})
        self.assertEquals(response.status_code, 302)
        prefix = f"{settings.REST_DOMAIN}/api/"
        posted_map = defaultdict(list)
        for url, data in self.posted:
            posted_map[url].append(json.loads(data))
        self.assertEquals(
            posted_map[prefix + "afsender"],
            [
                {
                    "navn": "TestFirma1",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "modtager"],
            [
                {
                    "navn": "TestFirma2",
                    "adresse": "Testvej 43",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "124",
                    "telefon": "123123",
                    "cvr": 12345679,
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "fragtforsendelse"],
            [
                {
                    "fragtbrevsnummer": "ABCDE1234567",
                    "forsendelsestype": "S",
                    "forbindelsesnr": "ABC 337",
                    "fragtbrev": base64.b64encode("Testtekst".encode("utf-8")).decode(
                        "ascii"
                    ),
                    "fragtbrev_navn": "fragtbrev.txt",
                    "afgangsdato": "2023-11-03",
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "afgiftsanmeldelse"],
            [
                {
                    "leverandørfaktura_nummer": "123",
                    "indførselstilladelse": "123",
                    "afsender_id": 1,
                    "modtager_id": 1,
                    "fuldmagtshaver_id": None,
                    "postforsendelse_id": None,
                    "fragtforsendelse_id": 1,
                    "leverandørfaktura": base64.b64encode(
                        "Testtekst".encode("utf-8")
                    ).decode("ascii"),
                    "leverandørfaktura_navn": "leverandørfaktura.txt",
                    "betales_af": "afsender",
                    "oprettet_på_vegne_af_id": 1,
                    "toldkategori": None,
                    "kladde": False,
                    "status": None,
                }
            ],
        )

    @patch.object(requests.Session, "get")
    @patch.object(requests.Session, "post")
    def test_form_successful_duplicate(self, mock_post, mock_get):
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get
        mock_post.side_effect = self.mock_requests_post
        response = self.client.post(url, data={**self.formdata1, **self.formfiles1})
        response = self.client.post(url, data={**self.formdata1, **self.formfiles1})
        self.assertEquals(response.status_code, 302)
        prefix = f"{settings.REST_DOMAIN}/api/"
        posted_map = defaultdict(list)
        for url, data in self.posted:
            posted_map[url].append(json.loads(data))
        self.assertEquals(len(posted_map[prefix + "afsender"]), 1)
        self.assertEquals(len(posted_map[prefix + "modtager"]), 1)
        self.assertEquals(len(posted_map[prefix + "fragtforsendelse"]), 1)
        self.assertEquals(len(posted_map[prefix + "afgiftsanmeldelse"]), 1)

    @patch.object(requests.Session, "get")
    @patch.object(requests.Session, "post")
    def test_form_successful_preexisting_actors(self, mock_post, mock_get):
        self.mock_existing["afsender"] = True
        self.mock_existing["modtager"] = True
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get
        mock_post.side_effect = self.mock_requests_post
        response = self.client.post(url, data={**self.formdata1, **self.formfiles1})
        self.assertEquals(response.status_code, 302, self.get_errors(response.content))
        prefix = f"{settings.REST_DOMAIN}/api/"
        posted_map = defaultdict(list)
        for url, data in self.posted:
            posted_map[url].append(json.loads(data))
        self.assertEquals(
            posted_map[prefix + "afsender"],
            [],
        )
        self.assertEquals(
            posted_map[prefix + "modtager"],
            [],
        )
        self.assertEquals(
            posted_map[prefix + "fragtforsendelse"],
            [
                {
                    "fragtbrevsnummer": "ABCDE1234567",
                    "forbindelsesnr": "ABC 337",
                    "forsendelsestype": "S",
                    "fragtbrev": base64.b64encode("Testtekst".encode("utf-8")).decode(
                        "ascii"
                    ),
                    "fragtbrev_navn": "fragtbrev.txt",
                    "afgangsdato": "2023-11-03",
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "afgiftsanmeldelse"],
            [
                {
                    "leverandørfaktura_nummer": "123",
                    "indførselstilladelse": "123",
                    "afsender_id": 1,
                    "modtager_id": 1,
                    "fuldmagtshaver_id": None,
                    "postforsendelse_id": None,
                    "fragtforsendelse_id": 1,
                    "leverandørfaktura": base64.b64encode(
                        "Testtekst".encode("utf-8")
                    ).decode("ascii"),
                    "leverandørfaktura_navn": "leverandørfaktura.txt",
                    "betales_af": "afsender",
                    "oprettet_på_vegne_af_id": 1,
                    "toldkategori": None,
                    "kladde": False,
                    "status": None,
                }
            ],
        )

    @patch.object(requests.Session, "get")
    @patch.object(requests.Session, "post")
    def test_form_successful_postforsendelse(self, mock_post, mock_get):
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get
        mock_post.side_effect = self.mock_requests_post
        response = self.client.post(url, data={**self.formdata2, **self.formfiles2})
        self.assertEquals(response.status_code, 302)
        prefix = f"{settings.REST_DOMAIN}/api/"
        posted_map = defaultdict(list)
        for url, data in self.posted:
            posted_map[url].append(json.loads(data))
        self.assertEquals(
            posted_map[prefix + "afsender"],
            [
                {
                    "navn": "TestFirma1",
                    "adresse": "Testvej 42",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "123",
                    "telefon": "123456",
                    "cvr": 12345678,
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "modtager"],
            [
                {
                    "navn": "TestFirma2",
                    "adresse": "Testvej 43",
                    "postnummer": 1234,
                    "by": "TestBy",
                    "postbox": "124",
                    "telefon": "123123",
                    "cvr": 12345679,
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "postforsendelse"],
            [
                {
                    "postforsendelsesnummer": "ABCDE1234567",
                    "forsendelsestype": "F",
                    "afsenderbykode": "1337",
                    "afgangsdato": "2023-11-03",
                    "kladde": False,
                }
            ],
        )
        self.assertEquals(
            posted_map[prefix + "afgiftsanmeldelse"],
            [
                {
                    "leverandørfaktura_nummer": "123",
                    "indførselstilladelse": "123",
                    "afsender_id": 1,
                    "modtager_id": 1,
                    "fuldmagtshaver_id": None,
                    "postforsendelse_id": 1,
                    "fragtforsendelse_id": None,
                    "leverandørfaktura": base64.b64encode(
                        "Testtekst".encode("utf-8")
                    ).decode("ascii"),
                    "leverandørfaktura_navn": "leverandørfaktura.txt",
                    "betales_af": "afsender",
                    "oprettet_på_vegne_af_id": 1,
                    "toldkategori": None,
                    "kladde": False,
                    "status": None,
                }
            ],
        )

    @patch.object(requests.Session, "get")
    @patch("admin.views.TF10FormCreateView.form_valid")
    def test_form_filefields_size(self, mock_form_valid, mock_get):
        self.login()
        url = reverse("tf10_create")
        mock_get.side_effect = self.mock_requests_get
        mock_form_valid.return_value = redirect(reverse("tf10_create"))
        data = {**self.formdata1}
        files = {
            "fragtbrev": SimpleUploadedFile("fragtbrev.txt", b"\x00" * 11000000),
            "leverandørfaktura": SimpleUploadedFile(
                "leverandørfaktura.txt", b"\x00" * 11000000
            ),
        }
        for v in files.values():
            v.seek(0)
        form = TF10CreateForm(data=data, files=files, oprettet_på_vegne_af_choices=())
        html_errors = self.submit_get_errors(url, {**data, **files})
        for file_field in (
            "leverandørfaktura",
            "fragtbrev",
        ):
            self.assertTrue(file_field in form.errors)
            self.assertEquals(
                form.errors[file_field], ["Filen er for stor; den må max. være 10.0 MB"]
            )
            self.assertTrue(file_field in html_errors)
            self.assertEquals(
                html_errors[file_field], ["Filen er for stor; den må max. være 10.0 MB"]
            )


class StatistikTest(PermissionsTest, TestCase):
    view = StatistikView

    check_permissions = ((reverse("statistik"), view.required_permissions),)

    def setUp(self):
        super().setUp()
        self.posted: List[Tuple[str, str]] = []

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path_parts = path.split("?")
        path = path_parts[0]
        query = parse_qs(path_parts[1])
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "statistik":
            if query.get("anmeldelsestype") in (["tf10"], None):
                json_content = {
                    "count": 2,
                    "items": [
                        {
                            "vareafgiftssats": 2,
                            "sum_afgiftsbeløb": "1750.00",
                            "afgiftsgruppenummer": 1234,
                            "vareart_da": "Kaffe",
                            "vareart_kl": "Kaffe",
                            "enhed": "kg",
                            "sum_mængde": "1000.000",
                            "sum_antal": 0,
                        },
                        {
                            "vareafgiftssats": 1,
                            "sum_afgiftsbeløb": "1400000.00",
                            "afgiftsgruppenummer": 5678,
                            "vareart_da": "Te",
                            "vareart_kl": "Te",
                            "enhed": "kg",
                            "sum_mængde": "1000.000",
                            "sum_antal": 0,
                        },
                    ],
                }
            else:
                json_content = {"count": 0, "items": []}
        elif path == expected_prefix + "vareafgiftssats":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftstabel": 1,
                        "vareart_da": "Kaffe",
                        "vareart_kl": "Kaffe",
                        "afgiftsgruppenummer": 1234,
                        "enhed": "kg",
                        "afgiftssats": "1.00",
                    },
                    {
                        "id": 2,
                        "afgiftstabel": 1,
                        "vareart_da": "Te",
                        "vareart_kl": "Te",
                        "afgiftsgruppenummer": 5678,
                        "enhed": "kg",
                        "afgiftssats": "1.00",
                    },
                ],
            }
        else:
            print(f"Mock got unrecognized path: {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @staticmethod
    def get_html_list(html: str):
        soup = BeautifulSoup(html, "html.parser")
        tables_in = soup.css.select("table")
        tables_out = []
        for table in tables_in:
            headers = [element.text for element in table.css.select("thead tr th")]
            tables_out.append(
                [
                    dict(zip(headers, [td.text.strip() for td in row.select("td")]))
                    for row in table.css.select("table tbody tr")
                ]
            )
        return tables_out

    @patch.object(requests.Session, "get")
    def test_submit_basic(self, mock_get):
        self.login()
        url = reverse("statistik")
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(
            url,
            data={
                "form-TOTAL_FORMS": 0,
                "form-INITIAL_FORMS": 0,
                "form-MIN_NUM_FORMS": 0,
                "form-MAX_NUM_FORMS": 1000,
            },
        )
        self.assertTrue(isinstance(response, TemplateResponse))
        tabledata = self.get_html_list(response.content)
        summary_table = tabledata[0]
        self.assertEquals(
            summary_table,
            [
                {
                    "Afgiftsgruppe": "1234",
                    "Afgiftstekst": "Kaffe",
                    "Afgift": "1.750,00",
                },
                {
                    "Afgiftsgruppe": "5678",
                    "Afgiftstekst": "Te",
                    "Afgift": "1.400.000,00",
                },
            ],
        )
        group_table = tabledata[1]
        self.assertEquals(
            group_table,
            [
                {"Gruppe": "Ingen data"},
            ],
        )

    @patch.object(requests.Session, "get")
    def test_submit_filter_1(self, mock_get):
        self.login()
        url = reverse("statistik")
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(
            url,
            data={
                "anmeldelsestype": "tf10",
                "form-TOTAL_FORMS": 0,
                "form-INITIAL_FORMS": 0,
                "form-MIN_NUM_FORMS": 0,
                "form-MAX_NUM_FORMS": 1000,
            },
        )
        self.assertTrue(isinstance(response, TemplateResponse))
        tabledata = self.get_html_list(response.content)
        summary_table = tabledata[0]
        self.assertEquals(
            summary_table,
            [
                {
                    "Afgiftsgruppe": "1234",
                    "Afgiftstekst": "Kaffe",
                    "Afgift": "1.750,00",
                },
                {
                    "Afgiftsgruppe": "5678",
                    "Afgiftstekst": "Te",
                    "Afgift": "1.400.000,00",
                },
            ],
        )
        group_table = tabledata[1]
        self.assertEquals(
            group_table,
            [
                {"Gruppe": "Ingen data"},
            ],
        )

    @patch.object(requests.Session, "get")
    def test_submit_filter_2(self, mock_get):
        self.login()
        url = reverse("statistik")
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(
            url,
            data={
                "anmeldelsestype": "tf5",
                "form-TOTAL_FORMS": 0,
                "form-INITIAL_FORMS": 0,
                "form-MIN_NUM_FORMS": 0,
                "form-MAX_NUM_FORMS": 1000,
            },
        )
        self.assertTrue(isinstance(response, TemplateResponse))
        tabledata = self.get_html_list(response.content)
        summary_table = tabledata[0]
        self.assertEquals(
            summary_table,
            [
                {"Afgiftsgruppe": "Ingen data"},
            ],
        )
        group_table = tabledata[1]
        self.assertEquals(
            group_table,
            [
                {"Gruppe": "Ingen data"},
            ],
        )

    @patch.object(requests.Session, "get")
    def test_submit_group(self, mock_get):
        self.login()
        url = reverse("statistik")
        mock_get.side_effect = self.mock_requests_get
        response = self.client.post(
            url,
            data={
                "form-TOTAL_FORMS": 1,
                "form-INITIAL_FORMS": 0,
                "form-MIN_NUM_FORMS": 0,
                "form-MAX_NUM_FORMS": 1000,
                "form-0-gruppe": "1234+5678",
            },
        )
        self.assertTrue(isinstance(response, TemplateResponse))
        tabledata = self.get_html_list(response.content)
        summary_table = tabledata[0]
        self.assertEquals(
            summary_table,
            [
                {
                    "Afgift": "1.750,00",
                    "Afgiftsgruppe": "1234",
                    "Afgiftstekst": "Kaffe",
                },
                {
                    "Afgift": "1.400.000,00",
                    "Afgiftsgruppe": "5678",
                    "Afgiftstekst": "Te",
                },
            ],
        )
        group_table = tabledata[1]
        self.assertEquals(
            group_table, [{"Afgift": "1.401.750,00", "Gruppe": "1234+5678"}]
        )


class TF5NotatTest(HasLogin, TestCase):
    formdata = {
        "cpr": "1234567890",
        "navn": "TestPerson1",
        "adresse": "Testvej 42",
        "postnummer": "1234",
        "by": "TestBy",
        "telefon": "123456",
        "bookingnummer": "123",
        "indførselstilladelse": "123",
        "leverandørfaktura_nummer": "123",
        "indleveringsdato": "2023-11-03",
        "form-TOTAL_FORMS": "1",
        "form-INITIAL_FORMS": "1",
        "form-0-id": "1",
        "form-0-vareafgiftssats": "1",
        "form-0-mængde": "1.00",
        "form-0-antal": "5",
        "form-0-fakturabeløb": "25000.00",
        "notat": "Dette er en test",
    }

    _formfiles1 = {
        "leverandørfaktura": SimpleUploadedFile(
            "leverandørfaktura.txt", "Testtekst".encode("utf-8")
        ),
    }

    def setUp(self):
        super().setUp()
        self.patched: List[Tuple[str, str]] = []
        self.posted: List[Tuple[str, str]] = []

    def mock_requests_post(self, path, data, headers=None):
        path = path.rstrip("/")
        response = Response()
        content = None
        status_code = None
        json_content = {"id": 1}
        self.posted.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_patch(self, path, data, headers=None):
        path = path.rstrip("/")
        response = Response()
        content = None
        status_code = None
        json_content = {"id": 1}
        self.patched.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "gyldig_fra": "2023-01-01T00:00:00-03:00",
                        "gyldig_til": None,
                        "kladde": False,
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftstabel": 1,
                        "vareart_da": "Båthorn",
                        "vareart_kl": "Båthorn",
                        "afgiftsgruppenummer": 1234567,
                        "enhed": "kg",
                        "afgiftssats": "1.00",
                        "kræver_indførselstilladelse": False,
                        "har_privat_tillægsafgift_alkohol": False,
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats/1":
            json_content = {
                "id": 1,
                "afgiftstabel": 1,
                "vareart_da": "Båthorn",
                "vareart_kl": "Båthorn",
                "afgiftsgruppenummer": 1234567,
                "enhed": "kg",
                "afgiftssats": "1.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            }
        elif path == expected_prefix + "varelinje":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "vareafgiftssats": 1,
                        "antal": 5,
                        "mængde": "1.00",
                        "fakturabeløb": "25000.00",
                        "afgiftsbeløb": "5000.00",
                    }
                ],
            }
        elif path == expected_prefix + "privat_afgiftsanmeldelse/1":
            json_content = {
                "id": 1,
                "cpr": 111111111,
                "navn": "Dummybruger Testersen",
                "adresse": "Testvej 12",
                "postnummer": 1234,
                "by": "Testby",
                "telefon": "123456",
                "bookingnummer": "1234",
                "indleveringsdato": "2024-02-01",
                "leverandørfaktura_nummer": "1234",
                "indførselstilladelse": None,
                "leverandørfaktura": None,  # "/privatfakturaer/1/a.pdf",
                "oprettet": "2024-01-24T12:23:17.315148+00:00",
                "oprettet_af": {
                    "id": 10,
                    "username": "0111111111 / 12345678",
                    "first_name": "Dummybruger",
                    "last_name": "Testersen",
                    "email": "test@magenta.dk",
                    "is_superuser": False,
                    "groups": ["PrivatIndberettere", "ErhvervIndberettere"],
                    "permissions": [
                        "aktør.add_afsender",
                        "aktør.add_modtager",
                        "aktør.change_afsender",
                        "aktør.change_modtager",
                        "aktør.view_afsender",
                        "aktør.view_modtager",
                        "anmeldelse.add_afgiftsanmeldelse",
                        "anmeldelse.add_notat",
                        "anmeldelse.add_privatafgiftsanmeldelse",
                        "anmeldelse.add_varelinje",
                        "anmeldelse.change_afgiftsanmeldelse",
                        "anmeldelse.change_privatafgiftsanmeldelse",
                        "anmeldelse.change_varelinje",
                        "anmeldelse.view_afgiftsanmeldelse",
                        "anmeldelse.view_notat",
                        "anmeldelse.view_privatafgiftsanmeldelse",
                        "anmeldelse.view_varelinje",
                        "forsendelse.add_fragtforsendelse",
                        "forsendelse.add_postforsendelse",
                        "forsendelse.change_fragtforsendelse",
                        "forsendelse.change_postforsendelse",
                        "forsendelse.delete_fragtforsendelse",
                        "forsendelse.delete_postforsendelse",
                        "forsendelse.view_fragtforsendelse",
                        "forsendelse.view_postforsendelse",
                        "payment.add_item",
                        "payment.add_payment",
                        "payment.change_item",
                        "payment.change_payment",
                        "payment.view_item",
                        "payment.view_payment",
                        "sats.view_afgiftstabel",
                        "sats.view_vareafgiftssats",
                    ],
                    "indberetter_data": {"cpr": 111111111, "cvr": 12345678},
                },
                "status": "ny",
                "anonym": False,
                "payment_status": "created",
            }

        else:
            print(f"Mock got unrecognized path: {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests.sessions.Session, "patch")
    @patch.object(requests.sessions.Session, "post")
    def test_add_notat(self, mock_post, mock_patch, mock_get):
        self.login()
        mock_get.side_effect = self.mock_requests_get
        mock_patch.side_effect = self.mock_requests_patch
        mock_post.side_effect = self.mock_requests_post
        response = self.client.post(
            reverse("tf5_edit", kwargs={"id": 1}),
            {**self.formdata, **self._formfiles1},
        )
        self.assertIn(
            (
                f"{settings.REST_DOMAIN}/api/notat",
                '{"afgiftsanmeldelse_id": null, "privatafgiftsanmeldelse_id": 1, "tekst": "Dette er en test"}',
            ),
            self.posted,
        )


class TF5BankPaymentTest(HasLogin, TestCase):
    view = TF5View

    check_permissions = (
        (reverse("tf5_view", kwargs={"id": 1}), view.required_permissions),
    )

    def setUp(self):
        super().setUp()
        self.posted: List[Tuple[str, str]] = []

    def mock_requests_post(self, path, data, headers=None):
        path = path.rstrip("/")
        response = Response()
        content = None
        status_code = None
        json_content = {"id": 1}
        self.posted.append((path, data))
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    def mock_requests_get(self, path):
        expected_prefix = f"{settings.REST_DOMAIN}/api/"
        path = path.split("?")[0]
        path = path.rstrip("/")
        response = Response()
        json_content = None
        content = None
        status_code = None
        if path == expected_prefix + "afgiftstabel":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "gyldig_fra": "2023-01-01T00:00:00-03:00",
                        "gyldig_til": None,
                        "kladde": False,
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftstabel": 1,
                        "vareart_da": "Båthorn",
                        "vareart_kl": "Båthorn",
                        "afgiftsgruppenummer": 1234567,
                        "enhed": "kg",
                        "afgiftssats": "1.00",
                        "kræver_indførselstilladelse": False,
                        "har_privat_tillægsafgift_alkohol": False,
                    }
                ],
            }
        elif path == expected_prefix + "vareafgiftssats/1":
            json_content = {
                "id": 1,
                "afgiftstabel": 1,
                "vareart_da": "Båthorn",
                "vareart_kl": "Båthorn",
                "afgiftsgruppenummer": 1234567,
                "enhed": "kg",
                "afgiftssats": "1.00",
                "kræver_indførselstilladelse": False,
                "har_privat_tillægsafgift_alkohol": False,
            }
        elif path == expected_prefix + "varelinje":
            json_content = {
                "count": 1,
                "items": [
                    {
                        "id": 1,
                        "afgiftsanmeldelse": 1,
                        "vareafgiftssats": 1,
                        "antal": 5,
                        "mængde": "1.00",
                        "fakturabeløb": "25000.00",
                        "afgiftsbeløb": "5000.00",
                    }
                ],
            }
        elif path == expected_prefix + "privat_afgiftsanmeldelse/1":
            json_content = {
                "id": 1,
                "cpr": 111111111,
                "navn": "Dummybruger Testersen",
                "adresse": "Testvej 12",
                "postnummer": 1234,
                "by": "Testby",
                "telefon": "123456",
                "bookingnummer": "1234",
                "indleveringsdato": "2024-02-01",
                "leverandørfaktura_nummer": "1234",
                "indførselstilladelse": None,
                "leverandørfaktura": None,  # "/privatfakturaer/1/a.pdf",
                "oprettet": "2024-01-24T12:23:17.315148+00:00",
                "oprettet_af": {
                    "id": 10,
                    "username": "0111111111 / 12345678",
                    "first_name": "Dummybruger",
                    "last_name": "Testersen",
                    "email": "test@magenta.dk",
                    "is_superuser": False,
                    "groups": ["PrivatIndberettere", "ErhvervIndberettere"],
                    "permissions": [
                        "aktør.add_afsender",
                        "aktør.add_modtager",
                        "aktør.change_afsender",
                        "aktør.change_modtager",
                        "aktør.view_afsender",
                        "aktør.view_modtager",
                        "anmeldelse.add_afgiftsanmeldelse",
                        "anmeldelse.add_notat",
                        "anmeldelse.add_privatafgiftsanmeldelse",
                        "anmeldelse.add_varelinje",
                        "anmeldelse.change_afgiftsanmeldelse",
                        "anmeldelse.change_privatafgiftsanmeldelse",
                        "anmeldelse.change_varelinje",
                        "anmeldelse.view_afgiftsanmeldelse",
                        "anmeldelse.view_notat",
                        "anmeldelse.view_privatafgiftsanmeldelse",
                        "anmeldelse.view_varelinje",
                        "forsendelse.add_fragtforsendelse",
                        "forsendelse.add_postforsendelse",
                        "forsendelse.change_fragtforsendelse",
                        "forsendelse.change_postforsendelse",
                        "forsendelse.delete_fragtforsendelse",
                        "forsendelse.delete_postforsendelse",
                        "forsendelse.view_fragtforsendelse",
                        "forsendelse.view_postforsendelse",
                        "payment.add_item",
                        "payment.add_payment",
                        "payment.change_item",
                        "payment.change_payment",
                        "payment.view_item",
                        "payment.view_payment",
                        "sats.view_afgiftstabel",
                        "sats.view_vareafgiftssats",
                    ],
                    "indberetter_data": {"cpr": 111111111, "cvr": 12345678},
                },
                "status": "ny",
                "anonym": False,
                "payment_status": "created",
            }
        elif path == expected_prefix + "notat":
            json_content = {
                "count": 0,
                "items": [],
            }

        else:
            print(f"Mock got unrecognized path: {path}")
        if json_content:
            content = json.dumps(json_content).encode("utf-8")
        if content:
            if not status_code:
                status_code = 200
            response._content = content
        response.status_code = status_code or 404
        return response

    @patch.object(requests.sessions.Session, "get")
    @patch.object(requests.sessions.Session, "post")
    def test_create_bank_payment(self, mock_post, mock_get):
        self.login()
        mock_get.side_effect = self.mock_requests_get
        mock_post.side_effect = self.mock_requests_post
        response = self.client.get("/admin/blanket/tf5/1")
        soup = BeautifulSoup(response.content, "html.parser")
        buttons = [element.text.strip() for element in soup.css.select("button")]
        self.assertIn("Opret betaling", buttons)

        self.client.post("/admin/blanket/tf5/1", {"betalt": "true"})
        prefix = f"{settings.REST_DOMAIN}/api/"
        posted_map = defaultdict(list)
        for url, data in self.posted:
            posted_map[url].append(json.loads(data))
        self.assertEquals(
            posted_map[prefix + "payment"],
            [{"declaration_id": 1, "provider": "bank"}],
        )
