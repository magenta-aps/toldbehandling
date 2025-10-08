from __future__ import annotations

import csv
from decimal import Context, Decimal
from io import StringIO
from typing import Any, Callable, Dict, Iterable, List, Optional, Union

from django.core.files.uploadedfile import UploadedFile
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from told_common.data import Vareafgiftssats, format_decimal, unformat_decimal


class SpreadsheetImportException(Exception):
    pass


class SpreadsheetExport:
    @staticmethod
    def render_xlsx(
        headers: Iterable[str],
        items: Iterable[Iterable],
        filename: str,
        column_widths: Optional[List[int]] = None,
    ) -> HttpResponse:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        if headers:
            ws.append(headers)
        for item in items:
            ws.append(item)
        if column_widths:
            for i, column_width in enumerate(column_widths, 1):
                if column_width is not None:
                    ws.column_dimensions[get_column_letter(i)].width = column_width
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename={}".format(filename)},
        )
        wb.save(response)
        return response

    @staticmethod
    def render_csv(
        headers: Iterable[str],
        items: Iterable[Iterable],
        filename: str,
    ) -> HttpResponse:
        response = HttpResponse(
            content_type="text/csv",
            headers={"Content-Disposition": "attachment; filename={}".format(filename)},
        )
        writer = csv.writer(response)
        writer.writerow(headers)
        for item in items:
            writer.writerow(item)
        return response


class VareafgiftssatsSpreadsheetUtil:
    @staticmethod
    def spreadsheet_row(
        sats: Vareafgiftssats,
        headers: List[str],
        lookup_afgiftssats: Callable[[int], Vareafgiftssats],
    ) -> List:
        row = []
        for header in headers:
            value = getattr(sats, header, None)
            if value is not None:
                if header == "overordnet":
                    value = lookup_afgiftssats(value).afgiftsgruppenummer
                elif header in (
                    "afgiftssats",
                    "minimumsbeløb",
                    "segment_øvre",
                    "segment_nedre",
                ):
                    value = format_decimal(value)
                elif header in (
                    "kræver_indførselstilladelse_alkohol",
                    "kræver_indførselstilladelse_tobak",
                    "har_privat_tillægsafgift_alkohol",
                    "synlig_privat",
                ):
                    value = "ja" if value else "nej"
                elif header == "enhed":
                    value = value.name.lower()
            row.append(value)
        return row

    @staticmethod
    def parse_bool(value):
        return value.lower() in ("ja", "aap", "yes", "true", "1")

    header_definitions: list[dict] = [
        {
            "label": "Afgiftsgruppenummer",
            "field": "afgiftsgruppenummer",
            "parser": int,
            "required": True,
        },
        {
            "label": "Overordnet",
            "field": "overordnet",
            "parser": int,
            "required": False,
        },
        {
            "label": "Vareart (da)",
            "field": "vareart_da",
            "parser": str,
            "required": True,
        },
        {
            "label": "Vareart (kl)",
            "field": "vareart_kl",
            "parser": str,
            "required": True,
        },
        {
            "label": "Enhed",
            "field": "enhed",
            "parser": lambda value: getattr(Vareafgiftssats.Enhed, value.upper()).value,
            "required": True,
        },
        {
            "label": "Afgiftssats",
            "field": "afgiftssats",
            "parser": unformat_decimal,
            "required": False,
        },
        {
            "label": "Kræver indførselstilladelse - Alkohol",
            "field": "kræver_indførselstilladelse_alkohol",
            "parser": parse_bool,
            "required": True,
        },
        {
            "label": "Kræver indførselstilladelse - Tobak",
            "field": "kræver_indførselstilladelse_tobak",
            "parser": parse_bool,
            "required": True,
        },
        {
            "label": "Har privat tillægsafgift alkohol",
            "field": "har_privat_tillægsafgift_alkohol",
            "parser": parse_bool,
            "required": True,
        },
        {
            "label": "Synlig for private",
            "field": "synlig_privat",
            "parser": parse_bool,
            "required": True,
        },
        {
            "label": "Minimumsbeløb",
            "field": "minimumsbeløb",
            "parser": unformat_decimal,
            "required": False,
        },
        {
            "label": "Segment nedre",
            "field": "segment_nedre",
            "parser": unformat_decimal,
            "required": False,
        },
        {
            "label": "Segment øvre",
            "field": "segment_øvre",
            "parser": unformat_decimal,
            "required": False,
        },
    ]

    @staticmethod
    def get_header(label):
        for x in VareafgiftssatsSpreadsheetUtil.header_definitions:
            if x["label"] == label:
                return x

    @staticmethod
    def from_spreadsheet_row(headers: List[str], row: list) -> Dict[str, Any]:
        data = {}
        for i, label in enumerate(headers):
            if label is not None:
                header_obj = VareafgiftssatsSpreadsheetUtil.get_header(label)
                if header_obj:
                    try:
                        value = row[i]
                    except IndexError:
                        value = None
                    if value == "":
                        value = None
                    try:
                        if value is not None:
                            value = header_obj["parser"](value)
                    except (TypeError, ValueError) as e:
                        raise SpreadsheetImportException(
                            f"Fejl ved import af regneark: {e}"
                        )
                    data[header_obj["field"]] = value
        return data

    @staticmethod
    def load_csv(data: UploadedFile) -> List[dict]:
        d = data.read().decode("utf-8")
        reader = csv.reader(StringIO(d), delimiter=",", quotechar='"')
        headers = next(reader)
        VareafgiftssatsSpreadsheetUtil.validate_headers(headers)
        satser = []
        for row in reader:
            satser.append(
                VareafgiftssatsSpreadsheetUtil.from_spreadsheet_row(headers, row)
            )
        return satser

    @staticmethod
    def load_xlsx(data: UploadedFile) -> List[dict]:
        wb = load_workbook(data, data_only=True)
        sheet = wb.active
        values = sheet.values
        headers = next(values)
        VareafgiftssatsSpreadsheetUtil.validate_headers(headers)
        satser = []
        for row in values:
            satser.append(
                VareafgiftssatsSpreadsheetUtil.from_spreadsheet_row(headers, row)
            )
        return satser

    @staticmethod
    def validate_headers(headers):
        headers = set(headers)
        for header in VareafgiftssatsSpreadsheetUtil.header_definitions:
            label = header["label"]
            if label not in headers:
                raise SpreadsheetImportException(f"Mangler kolonne med {label}")

    @staticmethod
    def validate_satser(satser: List[Dict]):
        # Start enumeration ved 2 fordi rækkerne i regneark er
        # 1-indekserede og vi har en header-række

        # Tjek at felter har gyldige værdier
        for linje, vareafgiftssats in enumerate(satser, 2):
            for header in VareafgiftssatsSpreadsheetUtil.header_definitions:
                if header["required"] and vareafgiftssats[header["field"]] is None:
                    label = header["label"]
                    raise SpreadsheetImportException(
                        f'Mangler felt "{label}" på linje {linje}'
                    )

        # Tjek at afgiftsgruppenummer er unikt
        afgiftsgruppenumre = set()
        for linje, vareafgiftssats in enumerate(satser, 2):
            afgiftsgruppenummer = vareafgiftssats["afgiftsgruppenummer"]
            if afgiftsgruppenummer in afgiftsgruppenumre:
                linjenumre = [
                    # Start ved 2 fordi rækkerne i regneark er
                    # 1-indekserede og vi har en header-række
                    str(i)
                    for i, sats in enumerate(satser, 2)
                    if sats["afgiftsgruppenummer"] == afgiftsgruppenummer
                ]
                raise SpreadsheetImportException(
                    f"Afgiftsgruppenummer {afgiftsgruppenummer} optræder to gange "
                    f"(linjer: {', '.join(linjenumre)})"
                )
            afgiftsgruppenumre.add(vareafgiftssats["afgiftsgruppenummer"])

        by_afgiftsgruppenummer = {
            x["afgiftsgruppenummer"]: (x, i) for i, x in enumerate(satser, 2)
        }

        # Tjek at alle satser der peges på med "overordnet" eksisterer
        for linje, vareafgiftssats in enumerate(satser, 2):
            afgiftsgruppenummer = vareafgiftssats["afgiftsgruppenummer"]
            if (
                vareafgiftssats["overordnet"]
                and vareafgiftssats["overordnet"] not in afgiftsgruppenumre
            ):
                raise SpreadsheetImportException(
                    f"Afgiftssats med afgiftsgruppenummer "
                    f"{afgiftsgruppenummer} (linje {linje}) "
                    f"peger på overordnet "
                    f"{vareafgiftssats['overordnet']}, som ikke findes"
                )
            overordnet_id: int = vareafgiftssats["overordnet"]
            besøgt = set()
            while overordnet_id is not None:
                linje = by_afgiftsgruppenummer[afgiftsgruppenummer][1]
                if afgiftsgruppenummer == overordnet_id:
                    raise SpreadsheetImportException(
                        f"Vareafgiftssats {afgiftsgruppenummer} (linje {linje}) "
                        f"peger på sig selv som overordnet"
                    )
                if overordnet_id in besøgt:
                    overordnet_linje = by_afgiftsgruppenummer[overordnet_id][1]
                    raise SpreadsheetImportException(
                        f"Vareafgiftssats {afgiftsgruppenummer} (linje {linje}) "
                        f"har {overordnet_id} (linje {overordnet_linje}) "
                        f"som overordnet, men {overordnet_id} har også "
                        f"{afgiftsgruppenummer} i kæden af overordnede"
                    )
                besøgt.add(afgiftsgruppenummer)
                afgiftsgruppenummer = overordnet_id
                overordnet_id = by_afgiftsgruppenummer[afgiftsgruppenummer][0][
                    "overordnet"
                ]
